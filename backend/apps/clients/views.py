from rest_framework import viewsets, status, permissions, mixins
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from django.db.models import Q, Count
from django.contrib.auth import get_user_model
import logging
from .models import Client, ClientInteraction, ClientVisit, Appointment, FollowUp, Task, Announcement, Purchase, AuditLog, CustomerTag, CustomerInterest, CustomerImportAudit
from .serializers import (
    ClientSerializer, ClientInteractionSerializer, AppointmentSerializer, FollowUpSerializer, 
    TaskSerializer, AnnouncementSerializer, PurchaseSerializer, AuditLogSerializer,
    CustomerTagSerializer
)
from apps.users.permissions import IsRoleAllowed, CanDeleteCustomer
from apps.users.middleware import ScopedVisibilityMixin
from apps.core.mixins import GlobalDateFilterMixin
from apps.stores.models import Store
import csv
import io
import json
from datetime import datetime
from django.http import HttpResponse, StreamingHttpResponse
import re

logger = logging.getLogger(__name__)
from django.db import transaction
from django.db.utils import IntegrityError
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
# import openpyxl
# from openpyxl import Workbook

User = get_user_model()


class IsAdminOrManager(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and (getattr(request.user, 'role', None) in ['platform_admin', 'business_admin', 'manager'])

MAX_IMPORT_BATCH_SIZE = 10000
MAX_AUDIT_ERRORS_STORED = 500


def _validate_import_row(row_num, row, tenant, get_value):
    """Validate one import row. Returns ('valid', None) or ('invalid', err) or ('needs_attention', err)."""
    from shared.validators import normalize_phone_number
    name = get_value(row, ['Name', 'name', 'NAME', 'Customer Name', 'customer name', 'CUSTOMER NAME'])
    first_name = get_value(row, ['first_name', 'First Name', 'FIRST_NAME']) or (name.split(' ', 1)[0].strip() if name else '')
    last_name = get_value(row, ['last_name', 'Last Name', 'LAST_NAME']) or (name.split(' ', 1)[1].strip() if name and ' ' in name else '')
    if last_name and not first_name:
        first_name, last_name = last_name, ''
    phone_raw = get_value(row, ['phone', 'Phone', 'Mobile No', 'mobile no', 'Mobile'])
    phone = ''
    if phone_raw:
        try:
            if 'E+' in str(phone_raw).upper():
                phone = str(int(float(phone_raw)))
            else:
                phone = str(phone_raw).strip()
            phone = re.sub(r'\D', '', phone)
            if len(phone) == 11 and phone.startswith('0'):
                phone = phone[1:]
            elif phone.startswith('91') and len(phone) == 12:
                phone = phone[2:]
            elif len(phone) > 10:
                phone = phone[-10:]
        except (ValueError, TypeError):
            phone = re.sub(r'\D', '', str(phone_raw))
    email = get_value(row, ['email', 'Email', 'EMAIL']) or None
    if email and not email.strip():
        email = None
    if not email and not phone:
        return ('invalid', {'row': row_num, 'message': 'Either email or phone is required.', 'code': 'required_field', 'field': 'email, phone'})
    if phone:
        digits = re.sub(r'\D', '', phone)
        if not str(phone).strip().startswith('+'):
            if len(digits) == 10:
                phone = f'+91{digits}'
            elif len(digits) == 11 and digits.startswith('0'):
                phone = f'+91{digits[1:]}'
            elif digits.startswith('91') and len(digits) == 12:
                phone = f'+{digits}'
            else:
                phone = normalize_phone_number(phone) if len(digits) >= 7 else phone
        else:
            phone = normalize_phone_number(phone)
    existing = None
    if email:
        existing = Client.objects.filter(email=email, tenant=tenant, is_deleted=False).first()
    if not existing and phone:
        existing = Client.objects.filter(phone=phone, tenant=tenant, is_deleted=False).first()
    if existing:
        return ('already_exists', {'row': row_num, 'message': f'Customer with {email or phone} already exists.', 'code': 'already_exists', 'field': 'email or phone'})
    assigned_to_username = get_value(row, ['assigned_to', 'Assigned To', 'ASSIGNED_TO', 'Attended By', 'attended_by', 'Sales Person'])
    if assigned_to_username and assigned_to_username.strip():
        assigned_user = User.objects.filter(
            username=assigned_to_username.strip(),
            tenant=tenant
        ).first()
        if not assigned_user:
            return ('needs_attention', {
                'row': row_num,
                'message': f'Salesperson "{assigned_to_username}" not found. Please add, use "Import with name only", or "Auto-create salesperson".',
                'code': 'salesperson_not_found',
                'field': 'assigned_to',
            })
    return ('valid', None)


class ImportExportPermission(permissions.BasePermission):
    """
    Allows import/export operations only to business admins and managers.
    """
    def has_permission(self, request, view):
        user = request.user
        if not user or not user.is_authenticated:
            return False
        return user.role in ['platform_admin', 'business_admin', 'manager']


"""
CLIENT MANAGEMENT SYSTEM - CUSTOMER STATUS MANAGEMENT

IMPORTANT: Customer statuses are now simplified to three main categories:
VVIP, VIP, and GENERAL. This provides clear customer segmentation:

1. VVIP Customers: Highest value customers with premium status
2. VIP Customers: High-value customers with special privileges  
3. GENERAL Customers: Regular customers with standard service

All customers are visible in the main customer system with appropriate
status-based filtering and management capabilities.
"""

class ClientViewSet(viewsets.ModelViewSet, ScopedVisibilityMixin, GlobalDateFilterMixin):
    serializer_class = ClientSerializer
    permission_classes = [IsRoleAllowed.for_roles(['inhouse_sales','manager','business_admin'])]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    
    def get_serializer_context(self):
        """Override to pass import_created_at through context"""
        context = super().get_serializer_context()
        # Pass preserved created_at through context if available
        if hasattr(self, '_import_created_at') and self._import_created_at:
            context['import_created_at'] = self._import_created_at
        return context
    
    def get_permissions(self):
        """
        Override to use different permissions for different actions.
        """
        if self.action == 'destroy':
            return [CanDeleteCustomer()]
        return super().get_permissions()
    
    def get_queryset(self):
        """Filter clients by user scope and date range"""
        # For restore and permanent_delete actions, include soft-deleted clients
        if hasattr(self, 'action') and self.action in ['restore', 'permanent_delete']:
            queryset = self.get_scoped_queryset(Client)
        else:
            # Since we're doing hard delete, we don't need to filter by is_deleted
            queryset = self.get_scoped_queryset(Client)
        
        # Apply global date filtering
        queryset = self.get_date_filtered_queryset(queryset, 'created_at')
        
        return queryset
    
    def create(self, request, *args, **kwargs):
        try:
            # Extract created_at BEFORE validation (since serializer will remove it)
            # This is for historical imports - preserve the date
            import_created_at = request.data.get('created_at')
            if import_created_at:
                print(f"=== VIEW: Preserving created_at for import: {import_created_at} ===")
                # Store it in request.data temporarily, will be handled by serializer
                # But we need to preserve it since serializer removes it
                self._preserved_created_at = import_created_at
            else:
                self._preserved_created_at = None
            
            # Validate the data first
            serializer = self.get_serializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            # Preserve created_at from serializer if it was extracted
            if hasattr(serializer, '_import_created_at') and serializer._import_created_at:
                self._preserved_created_at = serializer._import_created_at
            
            # Set tenant and store automatically
            if request.user.tenant:
                request.data['tenant'] = request.user.tenant.id
            if request.user.store:
                request.data['store'] = request.user.store.id
            
            # Re-add created_at to request.data so it's available for the second serializer call
            # Also store it in view instance so we can pass it via context
            if self._preserved_created_at:
                # Try to modify request.data (may be QueryDict which is immutable)
                try:
                    if hasattr(request.data, '_mutable'):
                        request.data._mutable = True
                    request.data['created_at'] = self._preserved_created_at
                    if hasattr(request.data, '_mutable'):
                        request.data._mutable = False
                except Exception as e:
                    logger.warning("Could not modify request.data directly when setting created_at: %s", e)
                # Store in view instance so we can pass it via context
                self._import_created_at = self._preserved_created_at
                print(f"=== VIEW: Preserved created_at for second serializer call: {self._preserved_created_at} ===")
            
            # Set created_by to the selected salesperson, not the logged-in user
            # This allows multiple salespersons to use the same login account but be tracked individually
            selected_salesperson_id = request.data.get('sales_person_id')
            selected_salesperson_name = request.data.get('sales_person')
            assigned_to_value = request.data.get('assigned_to')  # This comes from CSV imports
            
            # Store the selected salesperson for use in perform_create
            self.selected_salesperson = None
            
            if selected_salesperson_id:
                # Use the provided salesperson ID
                try:
                    selected_user = User.objects.filter(
                        id=selected_salesperson_id,
                        role='inhouse_sales',
                        tenant=request.user.tenant,
                        is_active=True
                    ).first()
                    
                    if selected_user:
                        self.selected_salesperson = selected_user
                    else:
                        # Fallback to logged-in user if salesperson not found
                        self.selected_salesperson = request.user
                except Exception as e:
                    self.selected_salesperson = request.user
            elif selected_salesperson_name and selected_salesperson_name.strip() and selected_salesperson_name.lower() != 'not specified':
                # Try to find user by name (for backward compatibility)
                try:
                    # Try to find user by first_name and last_name combination
                    name_parts = selected_salesperson_name.split()
                    if len(name_parts) >= 2:
                        first_name = name_parts[0]
                        last_name = ' '.join(name_parts[1:])
                        selected_user = User.objects.filter(
                            first_name__iexact=first_name,
                            last_name__iexact=last_name,
                            role='inhouse_sales',
                            tenant=request.user.tenant,
                            is_active=True
                        ).first()
                    else:
                        # Try to find by username or first_name only
                        selected_user = User.objects.filter(
                            Q(username__iexact=selected_salesperson_name) |
                            Q(first_name__iexact=selected_salesperson_name),
                            role='inhouse_sales',
                            tenant=request.user.tenant,
                            is_active=True
                        ).first()
                    
                    if selected_user:
                        self.selected_salesperson = selected_user
                    else:
                        # Fallback to logged-in user if salesperson not found
                        self.selected_salesperson = request.user
                except Exception as e:
                    self.selected_salesperson = request.user
            elif assigned_to_value and assigned_to_value.strip() and assigned_to_value.lower() != 'not specified':
                # Check assigned_to field (this is what comes from CSV imports)
                # assigned_to should be a username string
                try:
                    assigned_to_username = assigned_to_value.strip()
                    # Try to find user by username (don't filter by role - CSV might have any user type)
                    selected_user = User.objects.filter(
                        username__iexact=assigned_to_username,
                        tenant=request.user.tenant,
                        is_active=True
                    ).first()
                    
                    if selected_user:
                        self.selected_salesperson = selected_user
                    else:
                        # Don't fallback to manager - let the serializer handle assigned_to lookup
                        # The serializer will handle the assigned_to field separately
                        # For created_by, we'll use the manager (request.user) since the user from CSV wasn't found
                        # Set to manager for created_by, but sales_person and assigned_to will still use CSV value
                        self.selected_salesperson = request.user
                except Exception as e:
                    self.selected_salesperson = request.user
            else:
                # No salesperson selected, use logged-in user
                self.selected_salesperson = request.user
            
            # Check for duplicate phone before creating - if found, update existing customer
            existing_phone_customer = None
            phone = request.data.get('phone')
            if phone:
                from shared.validators import normalize_phone_number
                normalized_phone = normalize_phone_number(phone)
                existing_phone_customer = Client.objects.filter(
                    phone=normalized_phone,
                    tenant=request.user.tenant,
                    is_deleted=False
                ).first()
            
            # If existing customer found, update instead of creating new
            if existing_phone_customer:
                # Update existing customer with new data
                serializer = self.get_serializer(existing_phone_customer, data=request.data, partial=True)
                if not serializer.is_valid():
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
                # Save the updated customer
                serializer.save()
                client = serializer.instance
                
                # Update the updated_at timestamp to current time (so it shows in current month)
                from django.utils import timezone
                client.updated_at = timezone.now()
                client.save(update_fields=['updated_at'])
                
                # Also update the SalesPipeline entry's updated_at so it shows in current month pipeline
                from apps.sales.models import SalesPipeline
                pipeline_entries = SalesPipeline.objects.filter(client=client)
                if pipeline_entries.exists():
                    pipeline_entries.update(updated_at=timezone.now())
                
                # Return updated customer data
                response_data = serializer.data
                response_data['updated'] = True
                response_data['message'] = 'Existing customer updated successfully'
                return Response(response_data, status=status.HTTP_200_OK)
            
            # No existing customer found, create new one
            response = super().create(request, *args, **kwargs)
            
            # Create appointment if follow-up date is provided
            if response.status_code == 201 and response.data:
                client_data = response.data
                next_follow_up = request.data.get('next_follow_up')
                
                # Debug statements removed for production
                
                if next_follow_up:
                    try:
                        from datetime import datetime
                        from .models import Appointment
                        
                        # Parse the follow-up date
                        follow_up_date = datetime.strptime(next_follow_up, '%Y-%m-%d').date()
                        
                        # Get custom time or use default
                        next_follow_up_time = request.data.get('next_follow_up_time', '10:00')
                        follow_up_time = datetime.strptime(next_follow_up_time, '%H:%M').time()
                        
                        # Debug statements removed for production
                        
                        # Create appointment for the follow-up
                        # Use the same salesperson for appointment assignment
                        appointment_created_by = request.user
                        appointment_assigned_to = request.user
                        
                        # If we found a selected salesperson, use them for the appointment too
                        if selected_salesperson_id and 'selected_user' in locals() and selected_user:
                            appointment_created_by = selected_user
                            appointment_assigned_to = selected_user
                        
                        # Build a clean customer name without 'None' when last name is missing
                        first = (client_data.get('first_name') or '').strip()
                        last = (client_data.get('last_name') or '').strip()
                        clean_name = (f"{first} {last}" if last else first).strip() or 'customer'

                        appointment_data = {
                            'client_id': client_data['id'],
                            'tenant': request.user.tenant,
                            'date': follow_up_date,
                            'time': follow_up_time,
                            'purpose': f"Follow-up for {clean_name}",
                            'notes': f"Follow-up appointment created automatically when customer was added. Summary: {request.data.get('summary_notes', 'No notes provided')}",
                            'status': 'scheduled',
                            'created_by': appointment_created_by,
                            'assigned_to': appointment_assigned_to,
                            'duration': 60,  # Default 1 hour
                            'requires_follow_up': False,  # This is the follow-up itself
                        }
                        
                        # Debug statements removed for production
                        
                        appointment = Appointment.objects.create(**appointment_data)
                        
                        # Debug statements removed for production
                        
                    except Exception as appointment_error:
                        # Debug statements removed for production
                        # Don't fail the customer creation if appointment creation fails
                        pass
                # Debug statements removed for production
            
            return response
        except Exception as e:
            # Debug statements removed for production
            return Response(
                {"error": str(e), "detail": "Internal server error"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    
    @action(detail=False, methods=['get'])
    def check_phone(self, request):
        """Check if a phone number already exists (for duplicate detection before creation)"""
        phone = request.query_params.get('phone')
        if not phone:
            return Response(
                {'error': 'Phone parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from shared.validators import normalize_phone_number
        normalized_phone = normalize_phone_number(phone)
        
        existing_customer = Client.objects.filter(
            phone=normalized_phone,
            tenant=request.user.tenant,
            is_deleted=False
        ).first()
        
        if existing_customer:
            # Get store name if available
            store_name = existing_customer.store.name if existing_customer.store else 'No store assigned'
            # Get current user's store
            current_user_store_id = request.user.store.id if request.user.store else None
            customer_store_id = existing_customer.store.id if existing_customer.store else None
            
            return Response({
                'exists': True,
                'customer': {
                    'id': existing_customer.id,
                    'name': existing_customer.full_name,
                    'email': existing_customer.email or 'No email',
                    'status': existing_customer.get_status_display(),
                    'phone': existing_customer.phone,
                    'total_visits': existing_customer.appointments.count(),
                    'last_visit': existing_customer.appointments.order_by('-date').first().date.isoformat() if existing_customer.appointments.exists() else None,
                    'store_name': store_name,
                    'store_id': customer_store_id
                },
                'is_different_store': current_user_store_id is not None and customer_store_id is not None and current_user_store_id != customer_store_id,
                'message': f'A customer with this phone number already exists in {store_name}. You can use the existing customer or create a new visit entry.'
            })
        
        return Response({
            'exists': False,
            'message': 'Phone number is available'
        })
    
    @action(detail=False, methods=['post'])
    def test(self, request):
        """Test endpoint to check if the API is working"""
        # Debug statements removed for production
        return Response({"message": "Test endpoint working", "data": request.data})
    
    def list(self, request, *args, **kwargs):
        """List clients with filtering support"""
        queryset = self.get_queryset()
        
        # Optimize query: prefetch pipelines and select_related for assigned_to/created_by (Assigned To column)
        queryset = queryset.select_related('assigned_to', 'created_by').prefetch_related('pipelines')
        
        # Apply search filter
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        
        # Apply status filter
        status = request.query_params.get('status')
        if status and status != 'all':
            queryset = queryset.filter(status=status)
        
        # Apply store filter
        store = request.query_params.get('store')
        if store and store != 'all':
            try:
                store_id = int(store)
                queryset = queryset.filter(store_id=store_id)
            except (ValueError, TypeError):
                # If store ID is invalid, ignore the filter
                pass
        
        # Apply date range filter
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if start_date and end_date:
            try:
                from datetime import datetime
                start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                queryset = queryset.filter(
                    created_at__gte=start_dt,
                    created_at__lte=end_dt
                )
            except ValueError:
                # If date parsing fails, ignore the filter
                pass
        
        # Apply pagination if needed
        page = request.query_params.get('page')
        if page:
            try:
                page_num = int(page)
                page_size = 50  # Default page size
                start = (page_num - 1) * page_size
                end = start + page_size
                queryset = queryset[start:end]
            except ValueError:
                pass
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_update(self, serializer):
        instance = self.get_object()
        # Pass user context to serializer for audit logging
        serializer.context['user'] = self.request.user
        instance._auditlog_user = self.request.user
        return serializer.save()

    def perform_create(self, serializer):
        """Override to set the correct created_by user and handle tenant/store assignment"""
        # Use the selected salesperson if available, otherwise use logged-in user
        if hasattr(self, 'selected_salesperson') and self.selected_salesperson:
            created_by_user = self.selected_salesperson
        else:
            created_by_user = self.request.user
        
        # Save the instance with the correct created_by user
        instance = serializer.save(created_by=created_by_user)
        
        # Set created_at if it was preserved for import (historical dates)
        # CRITICAL: This MUST happen before any save() calls to override auto_now_add
        parsed_created_at = None
        if hasattr(self, '_import_created_at') and self._import_created_at:
            try:
                from django.utils.dateparse import parse_datetime, parse_date
                from django.utils import timezone
                from datetime import datetime
                
                created_at_str = self._import_created_at
                # Strip any leading/trailing quotes (single or double) from date string
                created_at_str = created_at_str.strip().strip("'").strip('"').strip()
                
                # Parse the date - try DD-MM-YYYY FIRST (most common for imports)
                created_at = None
                
                # First try: DD-MM-YYYY format (Indian format - most common in CSV imports)
                try:
                    parsed_date = datetime.strptime(created_at_str.strip(), '%d-%m-%Y').date()
                    created_at = datetime.combine(parsed_date, datetime.min.time())
                except ValueError:
                    try:
                        # Try DD/MM/YYYY format
                        parsed_date = datetime.strptime(created_at_str.strip(), '%d/%m/%Y').date()
                        created_at = datetime.combine(parsed_date, datetime.min.time())
                    except ValueError:
                        # Try parse_date (for ISO format)
                        parsed_date = parse_date(created_at_str)
                        if parsed_date:
                            created_at = datetime.combine(parsed_date, datetime.min.time())
                        else:
                            # Try parse_datetime
                            created_at = parse_datetime(created_at_str)
                            if not created_at:
                                # Try other formats
                                for fmt in [
                                    '%Y-%m-%dT%H:%M:%S', 
                                    '%Y-%m-%d %H:%M:%S', 
                                    '%Y-%m-%d', 
                                    '%Y/%m/%d',
                                    '%d-%m-%Y %H:%M:%S',
                                ]:
                                    try:
                                        created_at = datetime.strptime(created_at_str.strip(), fmt)
                                        break
                                    except Exception:
                                        continue
                
                if created_at:
                    # Make timezone-aware if naive
                    if timezone.is_naive(created_at):
                        created_at = timezone.make_aware(created_at)
                    parsed_created_at = created_at
                else:
                    logger.warning("Could not parse created_at from import value '%s'", created_at_str)
            except Exception as e:
                logger.warning("Error parsing created_at '%s': %s", created_at_str, e)
        
        # Set tenant and store if not already set
        user = self.request.user
        needs_save = False
        update_fields_list = []
        
        if user.tenant and not instance.tenant:
            instance.tenant = user.tenant
            needs_save = True
            update_fields_list.append('tenant')
        if user.store and not instance.store:
            instance.store = user.store
            needs_save = True
            update_fields_list.append('store')
        
        # CRITICAL: Set created_at if we parsed it - MUST use update_fields to override auto_now_add
        if parsed_created_at:
            instance.created_at = parsed_created_at
            needs_save = True
            if 'created_at' not in update_fields_list:
                update_fields_list.append('created_at')
        
        # Set audit log user for tracking (before any save)
        instance._auditlog_user = created_by_user
        
        # CRITICAL: Save with update_fields to override auto_now_add for created_at
        # This MUST happen to preserve historical dates from CSV imports
        if needs_save:
            if update_fields_list:
                instance.save(update_fields=update_fields_list)
            else:
                instance.save()
        
        # Create notifications for new customer using the actual creator
        # Only create if this is a new customer (not an update)
        # Use a more robust check to prevent multiple calls
        if not hasattr(instance, '_notifications_created'):
            # Check database first to see if notifications already exist
            from apps.notifications.models import Notification
            from django.utils import timezone
            from datetime import timedelta
            
            recent_cutoff = timezone.now() - timedelta(seconds=60)
            existing = Notification.objects.filter(
                type='new_customer',
                metadata__customer_id=instance.id,
                created_at__gte=recent_cutoff
            ).exists()
            
            if not existing:
                self.create_customer_notifications(instance, created_by_user)
                instance._notifications_created = True
            else:
                print(f"⚠️  Notifications already exist in DB for customer {instance.id} - skipping create_customer_notifications call")
        
        return instance
    
    def create_customer_notifications(self, client, created_by_user):
        """Create notifications when a new customer is added."""
        try:
            from apps.notifications.models import Notification
            from apps.users.models import User
            from django.utils import timezone
            from datetime import timedelta
            from django.db import transaction
            
            print(f"=== CREATING CUSTOMER NOTIFICATIONS ===")
            print(f"Client: {client.first_name} {client.last_name} (ID: {client.id})")
            print(f"Client store: {client.store}")
            print(f"Client tenant: {client.tenant}")
            print(f"Created by user: {created_by_user.username} (role: {created_by_user.role})")
            print(f"Created by user tenant: {created_by_user.tenant}")
            print(f"Created by user store: {created_by_user.store}")
            
            # Use database transaction with SELECT FOR UPDATE to prevent race conditions
            with transaction.atomic():
                # Use select_for_update to lock the row and prevent concurrent execution
                # Check if notifications already exist for this customer creation (prevent duplicates)
                # Check for notifications created in the last 120 seconds for this customer
                recent_cutoff = timezone.now() - timedelta(seconds=120)
                existing_notifications = Notification.objects.filter(
                    type='new_customer',
                    metadata__customer_id=client.id,
                    created_at__gte=recent_cutoff
                ).select_for_update().count()
                
                if existing_notifications > 0:
                    print(f"⚠️  Notifications already exist ({existing_notifications}) for customer {client.id} - skipping to prevent duplicates")
                    return
                
                # Get all users who should receive notifications
                users_to_notify = []
                
                # The user who created the customer should get notified
                users_to_notify.append(created_by_user)
                print(f"Added creator: {created_by_user.username}")
                
                # Business admin should always get notified
                if created_by_user.tenant:
                    business_admins = User.objects.filter(
                        tenant=created_by_user.tenant,
                        role='business_admin',
                        is_active=True
                    )
                    print(f"Found {business_admins.count()} business admins: {[f'{admin.username} (active: {admin.is_active})' for admin in business_admins]}")
                    users_to_notify.extend(business_admins)
                
                # Determine which store to use for notifications
                # Use customer's store if available, otherwise use the sales person's store
                notification_store = client.store if client.store else created_by_user.store
                print(f"🔍 Notification store determination:")
                print(f"   Client store: {client.store.name if client.store else 'None'}")
                print(f"   Sales person store: {created_by_user.store.name if created_by_user.store else 'None'}")
                print(f"   Using store: {notification_store.name if notification_store else 'None'}")
                
                # Store users should get notified if customer is assigned to their store OR if sales person has a store
                if notification_store:
                    print(f"✅ Using store for notifications: {notification_store.name} (from {'customer' if client.store else 'sales person'})")
                    
                    # Store manager - notify manager of the store
                    # IMPORTANT: Use store_id instead of store object to ensure proper matching
                    store_managers = User.objects.filter(
                        tenant=created_by_user.tenant,
                        role='manager',
                        store_id=notification_store.id,
                        is_active=True
                    ).select_related('store', 'tenant')
                    
                    managers_info = [f'{manager.username} (store: {manager.store.name if manager.store else "None"}, store_id: {manager.store.id if manager.store else "None"})' for manager in store_managers]
                    print(f"🔍 Manager query details:")
                    print(f"   Tenant: {created_by_user.tenant} (ID: {created_by_user.tenant.id if created_by_user.tenant else 'None'})")
                    print(f"   Role: manager")
                    print(f"   Store ID: {notification_store.id}")
                    print(f"   Store Name: {notification_store.name}")
                    print(f"   Is Active: True")
                    print(f"✅ Found {store_managers.count()} store managers: {managers_info}")
                    
                    if store_managers.count() == 0:
                        # Debug: Check if any managers exist for this tenant at all
                        all_tenant_managers = User.objects.filter(
                            tenant=created_by_user.tenant,
                            role='manager',
                            is_active=True
                        ).select_related('store')
                        print(f"⚠️  WARNING: No managers found for store {notification_store.name}!")
                        print(f"   Total managers in tenant: {all_tenant_managers.count()}")
                        for mgr in all_tenant_managers:
                            print(f"   - {mgr.username} (store: {mgr.store.name if mgr.store else 'None'}, store_id: {mgr.store.id if mgr.store else 'None'})")
                    else:
                        users_to_notify.extend(store_managers)
                
                    # NOTE: We do NOT notify all inhouse_sales, telecalling, or marketing users
                    # Only the creator, business admin, and manager should be notified
                    # This prevents spam notifications to all store employees
                else:
                    print(f"⚠️  Neither customer nor sales person has a store assigned - only business admin will be notified")
                
                # Remove duplicates (in case created_by_user has multiple roles or is in multiple categories)
                unique_users = list({user.id: user for user in users_to_notify}.values())
                print(f"Total users to notify (before deduplication): {len(users_to_notify)}")
                print(f"Unique users to notify (after deduplication): {len(unique_users)}")
                
                # Create notifications for each user
                # Use atomic transaction to prevent duplicates
                notifications_created = 0
                
                for user in unique_users:
                    # Check if notification already exists for this user and customer (within last 120 seconds)
                    recent_cutoff = timezone.now() - timedelta(seconds=120)
                    
                    # Use select_for_update to lock and prevent duplicates
                    existing = Notification.objects.filter(
                        user=user,
                        tenant=client.tenant,
                        type='new_customer',
                        metadata__customer_id=client.id,
                        created_at__gte=recent_cutoff
                    ).select_for_update().first()
                    
                    if existing:
                        print(f"⚠️  Notification already exists (ID: {existing.id}) for user {user.username} and customer {client.id} - skipping")
                        continue
                    
                    # Create notification directly (we're already in a transaction with lock)
                    try:
                        # IMPORTANT: Ensure we're using the correct user object, not a stale reference
                        notification_user = user
                        print(f"🔔 Creating notification for user: {notification_user.id} ({notification_user.username}, role: {notification_user.role})")
                        
                        notification = Notification.objects.create(
                            user=notification_user,
                            tenant=client.tenant,
                            store=notification_store,
                            type='new_customer',
                            title='New customer registered',
                            message=f'{client.first_name} {client.last_name} has been registered as a new customer by {created_by_user.first_name or created_by_user.username}',
                            priority='medium',
                            status='unread',
                            action_url=f'/customers/{client.id}',
                            action_text='View Customer',
                            is_persistent=False,
                            metadata={'customer_id': client.id, 'created_by_user_id': created_by_user.id}
                        )
                        
                        # Verify the notification was created with the correct user
                        notification.refresh_from_db()
                        print(f"✅ Created notification {notification.id} for user {notification.user.id} ({notification.user.username}, role: {notification.user.role})")
                        print(f"   Notification details: user_id={notification.user.id}, tenant_id={notification.tenant.id}, store_id={notification.store.id if notification.store else None}")
                        
                        notifications_created += 1
                    except Exception as create_error:
                        print(f"❌ Error creating notification for user {user.username}: {create_error}")
                        # Continue with other users even if one fails
                        continue
                
                print(f"✅ Created {notifications_created} notifications for new customer {client.first_name} {client.last_name} (ID: {client.id})")
                print(f"Users notified: {[f'{user.username} ({user.role})' for user in unique_users]}")
            
        except Exception as e:
            print(f"Error creating notifications for new customer: {e}")
            import traceback
            print(f"Traceback: {traceback.format_exc()}")
            # Don't fail the customer creation if notification creation fails

    def update(self, request, *args, **kwargs):
        print(f"=== CLIENT VIEW UPDATE METHOD ===")
        print(f"Request data: {request.data}")
        print(f"Request method: {request.method}")
        print(f"Request user: {request.user}")
        
        try:
            response = super().update(request, *args, **kwargs)
            print(f"Update successful: {response.data}")
            return response
        except Exception as e:
            print(f"Update failed with error: {e}")
            print(f"Error type: {type(e)}")
            import traceback
            print(f"Traceback: {traceback.format_exc()}")
            raise

    def perform_destroy(self, instance):
        instance._auditlog_user = self.request.user
        # Hard delete - completely remove from database
        instance.delete()

    def destroy(self, request, *args, **kwargs):
        """
        Hard delete a client. Only business admins can delete customers.
        """
        try:
            instance = self.get_object()
            
            # Check if user has permission to delete this customer
            permission = CanDeleteCustomer()
            if not permission.has_object_permission(request, self, instance):
                return Response({
                    'error': 'You do not have permission to delete customers.',
                    'detail': 'Only business admins can delete customers.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Log the deletion before performing it
            logger.info(f"Deleting client: {instance.id} ({instance.first_name} {instance.last_name}) by user: {request.user.username}")
            
            # Create audit log before deletion
            instance._auditlog_user = request.user
            from .models import AuditLog
            before = {field.name: str(getattr(instance, field.name)) for field in instance._meta.fields}
            AuditLog.objects.create(
                client=instance,
                action='delete',
                user=request.user,
                before=before,
                after=None
            )
            
            # Perform hard delete
            instance.delete()
            
            logger.info(f"Successfully deleted client: {instance.id}")
            
            return Response({
                'success': True,
                'message': 'Customer permanently deleted from database'
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error deleting client: {e}")
            return Response({
                'error': 'Failed to delete customer',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='trash')
    def trash(self, request):
        """List all soft-deleted clients for the tenant."""
        queryset = Client.objects.filter(is_deleted=True)
        if request.user.is_authenticated:
            user_tenant = request.user.tenant
            if user_tenant:
                queryset = queryset.filter(tenant=user_tenant)
            else:
                queryset = Client.objects.none()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='restore')
    def restore(self, request, pk=None):
        client = self.get_object()
        if client.is_deleted:
            client.is_deleted = False
            client.deleted_at = None
            client._auditlog_user = request.user
            client.save()
            # Audit log for restore
            from .models import AuditLog
            AuditLog.objects.create(
                client=client,
                action='restore',
                user=request.user,
                before=None,
                after={field.name: str(getattr(client, field.name)) for field in client._meta.fields}
            )
            return Response({'status': 'client restored'})
        return Response({'error': 'client is not deleted'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['delete'], url_path='permanent')
    def permanent_delete(self, request, pk=None):
        client = self.get_object()
        if client.is_deleted:
            client._auditlog_user = request.user
            from .models import AuditLog
            before = {field.name: str(getattr(client, field.name)) for field in client._meta.fields}
            AuditLog.objects.create(
                client=client,
                action='delete',
                user=request.user,
                before=before,
                after=None
            )
            client.delete()
            return Response({'status': 'client permanently deleted'})
        return Response({'error': 'client must be soft-deleted first'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], permission_classes=[ImportExportPermission])
    def export_csv(self, request):
        """Export customers to CSV - only for business admin and managers"""
        try:
            search = request.query_params.get('search')
            exhibition_filter_param = request.query_params.get('exhibition')
            product_filter_param = request.query_params.get('product')

            # If search, exhibition, or product filter is present, use scoped queryset (bypasses default date filtering)
            if search or (exhibition_filter_param and exhibition_filter_param != 'all') or (product_filter_param and product_filter_param != 'all'):
                queryset = self.get_scoped_queryset(Client)
            else:
                queryset = self.get_queryset()  # Applies global date filtering

            # Apply date filters if provided (only if no search/exhibition, or if explicitly provided)
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            if start_date and end_date:
                try:
                    start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                    end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                    queryset = queryset.filter(
                        updated_at__gte=start_dt,
                        updated_at__lte=end_dt
                    )
                except ValueError:
                    pass

            # Apply search filter
            if search:
                search_terms = search.strip().split()
                if len(search_terms) >= 2:
                    first_name_search = search_terms[0]
                    last_name_search = ' '.join(search_terms[1:])
                    queryset = queryset.filter(
                        Q(first_name__icontains=first_name_search) & Q(last_name__icontains=last_name_search) |
                        Q(first_name__icontains=last_name_search) & Q(last_name__icontains=first_name_search) |
                        Q(first_name__icontains=search) |
                        Q(last_name__icontains=search) |
                        Q(email__icontains=search) |
                        Q(phone__icontains=search)
                    )
                else:
                    queryset = queryset.filter(
                        Q(first_name__icontains=search) |
                        Q(last_name__icontains=search) |
                        Q(email__icontains=search) |
                        Q(phone__icontains=search)
                    )

            # Apply status filter
            status_filter_value = request.query_params.get('status')
            if status_filter_value and status_filter_value != 'all':
                pipeline_filter = Q(pipelines__stage=status_filter_value)
                status_field_filter = Q(status=status_filter_value)
                queryset = queryset.filter(pipeline_filter | status_field_filter).distinct()

            # Apply store filter
            store = request.query_params.get('store')
            if store and store != 'all':
                try:
                    store_id = int(store)
                    queryset = queryset.filter(store_id=store_id)
                except (ValueError, TypeError):
                    pass

            # Apply lead source filter
            lead_source = request.query_params.get('lead_source')
            if lead_source and lead_source != 'all':
                queryset = queryset.filter(lead_source=lead_source)

            # Apply created_by filter
            created_by = request.query_params.get('created_by')
            if created_by and created_by != 'all':
                try:
                    created_by_id = int(created_by)
                    queryset = queryset.filter(created_by_id=created_by_id)
                except (ValueError, TypeError):
                    pass

            # Apply exhibition filter
            if exhibition_filter_param and exhibition_filter_param != 'all':
                try:
                    exhibition_id = int(exhibition_filter_param)
                    queryset = queryset.filter(exhibition_id=exhibition_id)
                except (ValueError, TypeError):
                    pass

            # Apply product interest filter
            product_filter_param = request.query_params.get('product')
            if product_filter_param and product_filter_param != 'all':
                try:
                    product_id = int(product_filter_param)
                    # Filter clients that have at least one interest with this product
                    queryset = queryset.filter(interests__product_id=product_id).distinct()
                except (ValueError, TypeError):
                    pass
            
            # Get requested fields from query parameters
            fields_param = request.GET.get('fields', '')
            if fields_param:
                requested_fields = fields_param.split(',')
            else:
                # Default fields if none specified
                requested_fields = [
                    'first_name', 'last_name', 'email', 'phone', 'customer_type',
                    'address', 'city', 'state', 'country', 'postal_code',
                    'date_of_birth', 'anniversary_date', 'preferred_metal', 'preferred_stone',
                    'ring_size', 'budget_range', 'lead_source', 'notes', 'community',
                    'mother_tongue', 'reason_for_visit', 'age_of_end_user', 'saving_scheme',
                    'catchment_area', 'next_follow_up', 'summary_notes', 'status',
                    'created_at', 'updated_at', 'tags'
                ]
            
            # Check if customer_interests is requested - if so, replace with product_name and category
            has_customer_interests = 'customer_interests' in requested_fields
            if has_customer_interests:
                requested_fields.remove('customer_interests')
                if 'product_name' not in requested_fields:
                    requested_fields.append('product_name')
                if 'category' not in requested_fields:
                    requested_fields.append('category')
            
            # Prefetch pipelines to avoid N+1 queries
            queryset = queryset.prefetch_related('pipelines')
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="customers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.DictWriter(response, fieldnames=requested_fields)
            writer.writeheader()
            
            for client in queryset:
                row = {}
                for field in requested_fields:
                    if field == 'date_of_birth' and client.date_of_birth:
                        row[field] = client.date_of_birth.strftime('%Y-%m-%d')
                    elif field == 'anniversary_date' and client.anniversary_date:
                        row[field] = client.anniversary_date.strftime('%Y-%m-%d')
                    elif field in ['created_at', 'updated_at']:
                        row[field] = getattr(client, field).strftime('%d-%m-%Y')
                    elif field == 'phone' and getattr(client, field):
                        row[field] = str(getattr(client, field))
                    elif field == 'created_by':
                        # Handle created_by field - show creator's name
                        if client.created_by:
                            created_by_name = f"{client.created_by.first_name or ''} {client.created_by.last_name or ''}".strip()
                            if not created_by_name:
                                created_by_name = client.created_by.username or ''
                            row[field] = created_by_name
                        else:
                            row[field] = ''
                    elif field == 'product_name':
                        # Extract product names from customer interests with purchase status
                        product_info = []
                        try:
                            for idx, interest in enumerate(client.interests.all(), 1):
                                if interest.product:
                                    product_name = interest.product.name
                                    category_name = interest.category.name if interest.category else 'N/A'
                                    purchase_status = ''
                                    if interest.is_purchased:
                                        purchase_status = ' (Purchased)'
                                    elif interest.is_not_purchased:
                                        purchase_status = ' (Not Purchased)'
                                    # Format: "Product Name (Customer Interest 1) - Category: Category Name [Purchase Status]"
                                    product_info.append(f"{product_name} (Customer Interest {idx}) - Category: {category_name}{purchase_status}")
                        except Exception as e:
                            print(f"Error exporting product_name: {e}")
                        row[field] = ' | '.join(product_info) if product_info else ''
                    elif field == 'category':
                        # Extract category names from customer interests with purchase status
                        category_info = []
                        try:
                            for idx, interest in enumerate(client.interests.all(), 1):
                                if interest.category:
                                    category_name = interest.category.name
                                    product_name = interest.product.name if interest.product else 'N/A'
                                    purchase_status = ''
                                    if interest.is_purchased:
                                        purchase_status = ' (Purchased)'
                                    elif interest.is_not_purchased:
                                        purchase_status = ' (Not Purchased)'
                                    # Format: "Category Name (Customer Interest 1) - Product: Product Name [Purchase Status]"
                                    category_info.append(f"{category_name} (Customer Interest {idx}) - Product: {product_name}{purchase_status}")
                        except Exception as e:
                            print(f"Error exporting category: {e}")
                        row[field] = ' | '.join(category_info) if category_info else ''
                    elif field == 'tags':
                        row[field] = ', '.join([tag.name for tag in client.tags.all()])
                    elif field == 'status':
                        # Get pipeline stage instead of generic status
                        pipeline_stage = None
                        try:
                            # Check if pipelines are already prefetched
                            if hasattr(client, '_prefetched_objects_cache') and 'pipelines' in client._prefetched_objects_cache:
                                pipelines = client._prefetched_objects_cache['pipelines']
                                if pipelines:
                                    latest_pipeline = max(pipelines, key=lambda p: p.updated_at)
                                    pipeline_stage = latest_pipeline.stage
                            else:
                                latest_pipeline = client.pipelines.order_by('-updated_at').first()
                                if latest_pipeline:
                                    pipeline_stage = latest_pipeline.stage
                        except Exception as e:
                            print(f"Error getting pipeline stage for export: {e}")
                        # Use pipeline stage if available, otherwise fall back to status
                        row[field] = pipeline_stage if pipeline_stage else (client.status or 'general')
                    else:
                        value = getattr(client, field, '')
                        row[field] = str(value) if value is not None else ''
                
                writer.writerow(row)
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Export failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], permission_classes=[ImportExportPermission])
    def export_json(self, request):
        """Export customers to JSON - only for business admin and managers"""
        try:
            search = request.query_params.get('search')
            exhibition_filter_param = request.query_params.get('exhibition')
            product_filter_param = request.query_params.get('product')

            # If search, exhibition, or product filter is present, use scoped queryset (bypasses default date filtering)
            if search or (exhibition_filter_param and exhibition_filter_param != 'all') or (product_filter_param and product_filter_param != 'all'):
                queryset = self.get_scoped_queryset(Client)
            else:
                queryset = self.get_queryset()  # Applies global date filtering

            # Apply date filters if provided (only if no search/exhibition, or if explicitly provided)
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            if start_date and end_date:
                try:
                    start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                    end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                    queryset = queryset.filter(
                        updated_at__gte=start_dt,
                        updated_at__lte=end_dt
                    )
                except ValueError:
                    pass

            # Apply search filter
            if search:
                search_terms = search.strip().split()
                if len(search_terms) >= 2:
                    first_name_search = search_terms[0]
                    last_name_search = ' '.join(search_terms[1:])
                    queryset = queryset.filter(
                        Q(first_name__icontains=first_name_search) & Q(last_name__icontains=last_name_search) |
                        Q(first_name__icontains=last_name_search) & Q(last_name__icontains=first_name_search) |
                        Q(first_name__icontains=search) |
                        Q(last_name__icontains=search) |
                        Q(email__icontains=search) |
                        Q(phone__icontains=search)
                    )
                else:
                    queryset = queryset.filter(
                        Q(first_name__icontains=search) |
                        Q(last_name__icontains=search) |
                        Q(email__icontains=search) |
                        Q(phone__icontains=search)
                    )

            # Apply status filter
            status_filter_value = request.query_params.get('status')
            if status_filter_value and status_filter_value != 'all':
                pipeline_filter = Q(pipelines__stage=status_filter_value)
                status_field_filter = Q(status=status_filter_value)
                queryset = queryset.filter(pipeline_filter | status_field_filter).distinct()

            # Apply store filter
            store = request.query_params.get('store')
            if store and store != 'all':
                try:
                    store_id = int(store)
                    queryset = queryset.filter(store_id=store_id)
                except (ValueError, TypeError):
                    pass

            # Apply lead source filter
            lead_source = request.query_params.get('lead_source')
            if lead_source and lead_source != 'all':
                queryset = queryset.filter(lead_source=lead_source)

            # Apply created_by filter
            created_by = request.query_params.get('created_by')
            if created_by and created_by != 'all':
                try:
                    created_by_id = int(created_by)
                    queryset = queryset.filter(created_by_id=created_by_id)
                except (ValueError, TypeError):
                    pass

            # Apply exhibition filter
            if exhibition_filter_param and exhibition_filter_param != 'all':
                try:
                    exhibition_id = int(exhibition_filter_param)
                    queryset = queryset.filter(exhibition_id=exhibition_id)
                except (ValueError, TypeError):
                    pass

            # Apply product interest filter
            product_filter_param = request.query_params.get('product')
            if product_filter_param and product_filter_param != 'all':
                try:
                    product_id = int(product_filter_param)
                    # Filter clients that have at least one interest with this product
                    queryset = queryset.filter(interests__product_id=product_id).distinct()
                except (ValueError, TypeError):
                    pass
            
            # Get requested fields from query parameters
            fields_param = request.GET.get('fields', '')
            if fields_param:
                requested_fields = fields_param.split(',')
            else:
                # Default fields if none specified
                requested_fields = [
                    'first_name', 'last_name', 'email', 'phone', 'customer_type',
                    'address', 'city', 'state', 'country', 'postal_code',
                    'date_of_birth', 'anniversary_date', 'preferred_metal', 'preferred_stone',
                    'ring_size', 'budget_range', 'lead_source', 'notes', 'community',
                    'mother_tongue', 'reason_for_visit', 'age_of_end_user', 'saving_scheme',
                    'catchment_area', 'next_follow_up', 'summary_notes', 'status',
                    'created_at', 'updated_at', 'tags'
                ]
            
            # Check if customer_interests is requested - if so, replace with product_name and category
            has_customer_interests = 'customer_interests' in requested_fields
            if has_customer_interests:
                requested_fields.remove('customer_interests')
                if 'product_name' not in requested_fields:
                    requested_fields.append('product_name')
                if 'category' not in requested_fields:
                    requested_fields.append('category')
            
            # Prefetch pipelines to avoid N+1 queries
            queryset = queryset.prefetch_related('pipelines')
            
            # Serialize data with only requested fields
            data = []
            for client in queryset:
                client_data = {}
                for field in requested_fields:
                    if field == 'date_of_birth' and client.date_of_birth:
                        client_data[field] = client.date_of_birth.strftime('%Y-%m-%d')
                    elif field == 'anniversary_date' and client.anniversary_date:
                        client_data[field] = client.anniversary_date.strftime('%Y-%m-%d')
                    elif field in ['created_at', 'updated_at']:
                        client_data[field] = getattr(client, field).strftime('%d-%m-%Y')
                    elif field == 'phone' and getattr(client, field):
                        client_data[field] = str(getattr(client, field))
                    elif field == 'created_by':
                        # Handle created_by field - show creator's name
                        if client.created_by:
                            created_by_name = f"{client.created_by.first_name or ''} {client.created_by.last_name or ''}".strip()
                            if not created_by_name:
                                created_by_name = client.created_by.username or ''
                            client_data[field] = created_by_name
                        else:
                            client_data[field] = ''
                    elif field == 'product_name':
                        # Extract product names from customer interests with purchase status
                        product_info = []
                        try:
                            for idx, interest in enumerate(client.interests.all(), 1):
                                if interest.product:
                                    product_name = interest.product.name
                                    category_name = interest.category.name if interest.category else 'N/A'
                                    purchase_status = ''
                                    if interest.is_purchased:
                                        purchase_status = ' (Purchased)'
                                    elif interest.is_not_purchased:
                                        purchase_status = ' (Not Purchased)'
                                    # Format: "Product Name (Customer Interest 1) - Category: Category Name [Purchase Status]"
                                    product_info.append(f"{product_name} (Customer Interest {idx}) - Category: {category_name}{purchase_status}")
                        except Exception as e:
                            print(f"Error exporting product_name: {e}")
                        client_data[field] = ' | '.join(product_info) if product_info else ''
                    elif field == 'category':
                        # Extract category names from customer interests with purchase status
                        category_info = []
                        try:
                            for idx, interest in enumerate(client.interests.all(), 1):
                                if interest.category:
                                    category_name = interest.category.name
                                    product_name = interest.product.name if interest.product else 'N/A'
                                    purchase_status = ''
                                    if interest.is_purchased:
                                        purchase_status = ' (Purchased)'
                                    elif interest.is_not_purchased:
                                        purchase_status = ' (Not Purchased)'
                                    # Format: "Category Name (Customer Interest 1) - Product: Product Name [Purchase Status]"
                                    category_info.append(f"{category_name} (Customer Interest {idx}) - Product: {product_name}{purchase_status}")
                        except Exception as e:
                            print(f"Error exporting category: {e}")
                        client_data[field] = ' | '.join(category_info) if category_info else ''
                    elif field == 'tags':
                        client_data[field] = [tag.name for tag in client.tags.all()]
                    elif field == 'status':
                        # Get pipeline stage instead of generic status
                        pipeline_stage = None
                        try:
                            # Check if pipelines are already prefetched
                            if hasattr(client, '_prefetched_objects_cache') and 'pipelines' in client._prefetched_objects_cache:
                                pipelines = client._prefetched_objects_cache['pipelines']
                                if pipelines:
                                    latest_pipeline = max(pipelines, key=lambda p: p.updated_at)
                                    pipeline_stage = latest_pipeline.stage
                            else:
                                latest_pipeline = client.pipelines.order_by('-updated_at').first()
                                if latest_pipeline:
                                    pipeline_stage = latest_pipeline.stage
                        except Exception as e:
                            print(f"Error getting pipeline stage for export: {e}")
                        # Use pipeline stage if available, otherwise fall back to status
                        client_data[field] = pipeline_stage if pipeline_stage else (client.status or 'general')
                    else:
                        value = getattr(client, field, '')
                        client_data[field] = value if value is not None else ''
                data.append(client_data)
            
            response = HttpResponse(
                json.dumps(data, indent=2, default=str),
                content_type='application/json'
            )
            response['Content-Disposition'] = f'attachment; filename="customers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Export failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    # @action(detail=False, methods=['get'], permission_classes=[ImportExportPermission])
    # def export_xlsx(self, request):
    #     """Export customers to XLSX - only for business admin and managers"""
    #     try:
    #         queryset = self.get_queryset()
    #         
    #         # Get requested fields from query parameters
    #         fields_param = request.GET.get('fields', '')
    #         if fields_param:
    #             requested_fields = fields_param.split(',')
    #         else:
    #             # Default fields if none specified
    #             requested_fields = [
    #                 'first_name', 'last_name', 'email', 'phone', 'customer_type',
    #                 'address', 'city', 'state', 'country', 'postal_code',
    #                 'date_of_birth', 'anniversary_date', 'preferred_metal', 'preferred_stone',
    #                 'ring_size', 'budget_range', 'lead_source', 'notes', 'community',
    #                 'mother_tongue', 'reason_for_visit', 'age_of_end_user', 'saving_scheme',
    #                 'catchment_area', 'next_follow_up', 'summary_notes', 'status',
    #                 'created_at', 'updated_at', 'tags'
    #             ]
    #         
    #         # Create Excel workbook
    #         wb = Workbook()
    #         ws = wb.active
    #         ws.title = "Customers"
    #         
    #         # Write headers
    #         for col, field in enumerate(requested_fields, 1):
    #             ws.cell(row=1, column=col, value=field.replace('_', ' ').title())
    #         
    #         # Write data
    #         for row, client in enumerate(queryset, 2):
    #             for col, field in enumerate(requested_fields, 1):
    #                 if field == 'date_of_birth' and client.date_of_birth:
    #                     value = client.date_of_birth.strftime('%Y-%m-%d')
    #                 elif field == 'anniversary_date' and client.anniversary_date:
    #                     value = client.anniversary_date.strftime('%Y-%m-%d')
    #                 elif field in ['created_at', 'updated_at']:
    #                     value = getattr(client, field).strftime('%Y-%m-%d %H:%M:%S')
    #                 elif field == 'tags':
    #                     value = ', '.join([tag.name for tag in client.tags.all()])
    #                 else:
    #                     value = getattr(client, field, '')
    #                     value = str(value) if value is not None else ''
    #                 
    #                 ws.cell(row=row, column=col, value=value)
    #             
    #             # Create response
    #             response = HttpResponse(
    #                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #             )
    #             response['Content-Disposition'] = f'attachment; filename="customers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    #             
    #             wb.save(response)
    #             return response
    #             
    #     except Exception as e:
    #         return Response(
    #                 {'error': f'Export failed: {str(e)}'}, 
    #                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
    #             )

    def _read_import_rows(self, file):
        """Read CSV or Excel file into list of row dicts. Returns (rows, error_response)."""
        if not any(file.name.lower().endswith(ext) for ext in ['.csv', '.xlsx', '.xls']):
            return None, Response({'error': 'Please upload a CSV, XLSX, or XLS file'}, status=status.HTTP_400_BAD_REQUEST)
        if file.name.lower().endswith('.csv'):
            try:
                decoded = file.read().decode('utf-8')
                rows = list(csv.DictReader(io.StringIO(decoded)))
            except Exception as e:
                return None, Response({'error': f'Error reading CSV: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            try:
                import openpyxl
                from openpyxl import load_workbook
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                headers = [cell.value or f'column_{i}' for i, cell in enumerate(ws[1])]
                rows = []
                for row_num in range(2, ws.max_row + 1):
                    row_data = {}
                    for col_num, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        row_data[header] = str(cell_value) if cell_value is not None else ''
                    rows.append(row_data)
            except ImportError:
                return None, Response({'error': 'Excel support requires openpyxl.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            except Exception as e:
                return None, Response({'error': f'Error reading Excel: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        return rows, None

    @action(detail=False, methods=['post'], permission_classes=[ImportExportPermission], parser_classes=[MultiPartParser, FormParser])
    def validate_import(self, request):
        """Validate customer import file and return report (no DB insert). Max 10,000 rows per file."""
        try:
            # DRF MultiPartParser: access request.data to trigger parsing; file may be in data or FILES
            data = getattr(request, 'data', None)
            file = (data.get('file') if data is not None else None) or request.FILES.get('file')
            if not file or not getattr(file, 'read', None):
                return Response({
                    'error': 'No file provided. Upload a CSV or Excel file using the "file" field.',
                }, status=status.HTTP_400_BAD_REQUEST)
            rows, err = self._read_import_rows(file)
            if err is not None:
                return err
            if not request.user.tenant:
                return Response({'error': 'User does not have a tenant assigned.'}, status=status.HTTP_400_BAD_REQUEST)
            # Business admin and platform admin can import without a store; others need a store
            if not request.user.store and request.user.role not in ('business_admin', 'platform_admin'):
                return Response({'error': 'User does not have a store assigned.'}, status=status.HTTP_400_BAD_REQUEST)
            if len(rows) > MAX_IMPORT_BATCH_SIZE:
                return Response({
                    'error': f'Maximum {MAX_IMPORT_BATCH_SIZE} customers per batch. Your file has {len(rows)} rows. If your dataset exceeds this limit, please split it into smaller files so each import contains no more than 10,000 rows.',
                }, status=status.HTTP_400_BAD_REQUEST)
            if len(rows) == 0:
                return Response({'error': 'No data rows found in file.'}, status=status.HTTP_400_BAD_REQUEST)

            stream = request.GET.get('stream') == '1' or request.query_params.get('stream') == '1'
            total_rows = len(rows)
            tenant = request.user.tenant

            def get_value(row, key_variations, default=''):
                for key in key_variations:
                    if key in row:
                        return (row[key] or '').strip() or default
                    for rk in row.keys():
                        if (rk or '').strip().lower() == key.lower().strip():
                            return (row[rk] or '').strip() or default
                return default

            errors = []
            valid_count = 0
            needs_attention_count = 0
            invalid_count = 0
            already_exists_count = 0

            if stream:
                def generate():
                    errs = []
                    v, i, n, a = 0, 0, 0, 0
                    yield f"data: {json.dumps({'type': 'progress', 'processed': 0, 'total': total_rows, 'valid_count': 0, 'invalid_count': 0, 'needs_attention_count': 0, 'already_exists_count': 0})}\n\n"
                    for row_num, row in enumerate(rows, start=2):
                        outcome, err = _validate_import_row(row_num, row, tenant, get_value)
                        if outcome == 'valid':
                            v += 1
                        elif outcome == 'invalid':
                            i += 1
                            errs.append(err)
                        elif outcome == 'already_exists':
                            a += 1
                            errs.append(err)
                        else:
                            n += 1
                            errs.append(err)
                        processed = row_num - 1
                        if processed % 100 == 0 or processed == total_rows:
                            yield f"data: {json.dumps({'type': 'progress', 'processed': processed, 'total': total_rows, 'valid_count': v, 'invalid_count': i, 'needs_attention_count': n, 'already_exists_count': a})}\n\n"
                    try:
                        errors_by_code = {}
                        for e in errs:
                            c = e.get('code') or 'unknown'
                            errors_by_code[c] = errors_by_code.get(c, 0) + 1
                        details = {
                            'errors_count': len(errs),
                            'errors_by_code': errors_by_code,
                            'errors': errs[:MAX_AUDIT_ERRORS_STORED],
                        }
                        CustomerImportAudit.objects.create(
                            user=request.user,
                            action='validated',
                            total_rows=total_rows,
                            valid_count=v,
                            invalid_count=i,
                            needs_attention_count=n,
                            details=details,
                        )
                    except Exception as audit_err:
                        logger.warning('Could not write CustomerImportAudit (run migrations?): %s', audit_err)
                    yield f"data: {json.dumps({'type': 'done', 'total_rows': total_rows, 'valid_count': v, 'invalid_count': i, 'needs_attention_count': n, 'already_exists_count': a, 'errors': errs, 'max_batch_size': MAX_IMPORT_BATCH_SIZE})}\n\n"

                return StreamingHttpResponse(
                    generate(),
                    content_type='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
                )

            for row_num, row in enumerate(rows, start=2):
                outcome, err = _validate_import_row(row_num, row, tenant, get_value)
                if outcome == 'valid':
                    valid_count += 1
                elif outcome == 'invalid':
                    invalid_count += 1
                    errors.append(err)
                elif outcome == 'already_exists':
                    already_exists_count += 1
                    errors.append(err)
                else:
                    needs_attention_count += 1
                    errors.append(err)

            total_rows = len(rows)
            try:
                errors_by_code = {}
                for e in errors:
                    c = e.get('code') or 'unknown'
                    errors_by_code[c] = errors_by_code.get(c, 0) + 1
                details = {
                    'errors_count': len(errors),
                    'errors_by_code': errors_by_code,
                    'errors': errors[:MAX_AUDIT_ERRORS_STORED],
                }
                CustomerImportAudit.objects.create(
                    user=request.user,
                    action='validated',
                    total_rows=total_rows,
                    valid_count=valid_count,
                    invalid_count=invalid_count,
                    needs_attention_count=needs_attention_count,
                    details=details,
                )
            except Exception as audit_err:
                logger.warning('Could not write CustomerImportAudit (run migrations?): %s', audit_err)
            return Response({
                'total_rows': total_rows,
                'valid_count': valid_count,
                'invalid_count': invalid_count,
                'needs_attention_count': needs_attention_count,
                'already_exists_count': already_exists_count,
                'errors': errors,
                'max_batch_size': MAX_IMPORT_BATCH_SIZE,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception('validate_import failed')
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], permission_classes=[ImportExportPermission])
    def import_file(self, request):
        """
        Import customers from CSV, XLSX, or XLS files - only for business admin and managers.

        NOTE: This endpoint now supports an optional `stream=1` query parameter to provide
        **server-sent event (SSE)** style progress updates while importing. When `stream=1`
        is present, the response will be a `text/event-stream` StreamingHttpResponse that
        sends `progress` and `done` events. Without `stream=1`, the original non-streaming
        JSON response is returned.
        """
        try:
            confirm = request.POST.get('confirm') == 'true' or request.data.get('confirm') is True
            if not confirm:
                return Response(
                    {'error': 'Import requires confirmation. Please review the validation report and confirm.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            salesperson_not_found = (request.POST.get('salesperson_not_found') or request.data.get('salesperson_not_found') or 'name_only').strip().lower()
            if salesperson_not_found not in ('skip', 'name_only', 'auto_create'):
                salesperson_not_found = 'name_only'

            if 'file' not in request.FILES:
                return Response(
                    {'error': 'No file provided'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            file = request.FILES['file']
            rows, read_err = self._read_import_rows(file)
            if read_err is not None:
                return read_err

            # Optional: only process specific row numbers (1-based file rows, e.g. [7474] for re-importing failed rows)
            only_rows_raw = request.POST.get('only_rows') or (request.data.get('only_rows') if hasattr(request, 'data') else None)
            if only_rows_raw:
                try:
                    if isinstance(only_rows_raw, list):
                        only_rows_set = {int(x) for x in only_rows_raw}
                    elif isinstance(only_rows_raw, str):
                        raw = only_rows_raw.strip()
                        if raw.startswith('['):
                            only_rows_set = {int(x) for x in json.loads(raw)}
                        else:
                            only_rows_set = {int(x.strip()) for x in raw.split(',') if x.strip()}
                    else:
                        only_rows_set = set()
                    if only_rows_set:
                        # File row 2 = rows[0], so row_num = index + 2
                        rows = [rows[i] for i in range(len(rows)) if (i + 2) in only_rows_set]
                        if not rows:
                            return Response(
                                {'error': 'No rows in file match the requested row numbers (only_rows).'},
                                status=status.HTTP_400_BAD_REQUEST
                            )
                except (ValueError, TypeError, json.JSONDecodeError) as e:
                    return Response(
                        {'error': f'Invalid only_rows: {str(e)}. Use a JSON array or comma-separated row numbers (e.g. [7474] or "7474").'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            if len(rows) > MAX_IMPORT_BATCH_SIZE:
                return Response({
                    'error': f'Maximum {MAX_IMPORT_BATCH_SIZE} customers per batch. Your file has {len(rows)} rows. If your dataset exceeds this limit, please split it into smaller files so each import contains no more than 10,000 rows.',
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if user has required assignments
            if not request.user.tenant:
                return Response(
                    {'error': 'User does not have a tenant assigned. Please contact your administrator.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Business admin and platform admin can import without a store; others need a store
            if not request.user.store and request.user.role not in ('business_admin', 'platform_admin'):
                return Response(
                    {'error': 'User does not have a store assigned. Please contact your administrator to assign you to a store.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Fallback store for admins who have no store (e.g. first store in tenant)
            import_store = request.user.store
            if not import_store and request.user.tenant:
                import_store = Store.objects.filter(tenant=request.user.tenant).first()

            # OPTIMIZATION: Pre-fetch existing customers, users, and stores to avoid per-row queries
            # Get all existing emails and phones in one query
            existing_emails = set(
                Client.objects.filter(tenant=request.user.tenant, is_deleted=False)
                .exclude(email__isnull=True).exclude(email='')
                .values_list('email', flat=True)
            )
            existing_phones = set(
                Client.objects.filter(tenant=request.user.tenant, is_deleted=False)
                .exclude(phone__isnull=True).exclude(phone='')
                .values_list('phone', flat=True)
            )
            # Maps for resolving existing customer so we can add ClientVisit instead of skipping
            email_to_client_id = dict(
                Client.objects.filter(tenant=request.user.tenant, is_deleted=False)
                .exclude(email__isnull=True).exclude(email='')
                .values_list('email', 'id')
            )
            phone_to_client_id = dict(
                Client.objects.filter(tenant=request.user.tenant, is_deleted=False)
                .exclude(phone__isnull=True).exclude(phone='')
                .values_list('phone', 'id')
            )
            
            # Pre-fetch all users in tenant for username lookups
            tenant_users = {u.username.lower(): u for u in User.objects.filter(tenant=request.user.tenant, is_active=True)}
            
            # Pre-fetch all stores in tenant
            tenant_stores = {s.name.lower(): s for s in Store.objects.filter(tenant=request.user.tenant)}
            # Also create normalized store lookup (remove punctuation)
            normalized_stores = {}
            for store_name, store in tenant_stores.items():
                normalized = re.sub(r'[^a-z0-9]', '', store_name)
                if normalized not in normalized_stores:
                    normalized_stores[normalized] = store
            
            # Shared import implementation so we can use it for both streaming and non-streaming modes
            def perform_import(stream_progress=False):
                # Use mutable dict so generator updates are visible to caller
                counters = {
                    'imported_count': 0,
                    'skipped_count': 0,
                    'visits_added_count': 0,  # ClientVisit created for existing customers (no row lost)
                    'errors': []
                }
                deferred_visits = []  # (email, phone, store_id, visit_date) for in-file duplicates until we have client_id
                clients_to_create = []  # Batch for bulk_create
                BATCH_SIZE = 500  # Process in batches of 500

                def parse_visit_from_row(row, get_value):
                    """Return (store_id, visit_date_dt) for this row; visit_date_dt is date or None."""
                    branch_name = get_value(['Branch', 'branch', 'BRANCH', 'Store', 'store', 'STORE'])
                    store_id = None
                    if branch_name and str(branch_name).strip():
                        bn = str(branch_name).strip().lower()
                        store = tenant_stores.get(bn)
                        if not store:
                            bn_normalized = re.sub(r'[^a-z0-9]', '', bn)
                            store = normalized_stores.get(bn_normalized)
                        if store:
                            store_id = store.id
                        elif import_store:
                            store_id = import_store.id
                    else:
                        store_id = import_store.id if import_store else None
                    created_at_str = None
                    for col_name in ['created_at', 'Created At', 'Created_at', 'CREATED_AT', 'Date', 'date', 'created_date', 'Created Date']:
                        if col_name in row or any(rk.lower().strip() == col_name.lower().strip() for rk in row.keys()):
                            val = get_value([col_name])
                            if val and str(val).strip():
                                created_at_str = str(val).strip()
                                break
                    visit_date_dt = None
                    if created_at_str:
                        try:
                            visit_date_dt = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S').date()
                        except ValueError:
                            try:
                                visit_date_dt = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M').date()
                            except ValueError:
                                try:
                                    visit_date_dt = datetime.strptime(created_at_str, '%Y-%m-%d').date()
                                except ValueError:
                                    try:
                                        visit_date_dt = datetime.strptime(created_at_str, '%d-%m-%Y').date()
                                    except ValueError:
                                        try:
                                            visit_date_dt = datetime.strptime(created_at_str, '%d/%m/%Y').date()
                                        except ValueError:
                                            pass
                    return (store_id, visit_date_dt)

                def row_iterator():
                    """
                    Generator that yields per-row import results and optionally progress events.
                    Each yielded value is a dict with:
                      - row_num
                      - imported_count, skipped_count, failed_count
                      - error (optional string)
                      - name (optional display name)
                    """
                    nonlocal clients_to_create

                    with transaction.atomic():
                        for row_num, row in enumerate(rows, start=2):  # Start from 2 to account for header
                            row_error = None
                            first_name = ''
                            last_name = ''
                            try:
                                # Helper function to get value with case-insensitive key matching
                                def get_value(key_variations, default=''):
                                    for key in key_variations:
                                        # Try exact match first
                                        if key in row:
                                            return row[key].strip() if row[key] else default
                                        # Try case-insensitive match
                                        for row_key in row.keys():
                                            if row_key.lower().strip() == key.lower().strip():
                                                return row[row_key].strip() if row[row_key] else default
                                    return default
                                
                                # Handle Name field - split into first_name and last_name
                                name = get_value(['Name', 'name', 'NAME', 'Customer Name', 'customer name', 'CUSTOMER NAME'])
                                first_name = ''
                                last_name = ''
                                if name and str(name).strip():
                                    name_parts = str(name).strip().split(' ', 1)
                                    first_name = name_parts[0].strip()
                                    last_name = name_parts[1].strip() if len(name_parts) > 1 else ''
                                if not first_name or not last_name:
                                    fn = get_value(['first_name', 'First Name', 'FIRST_NAME'])
                                    ln = get_value(['last_name', 'Last Name', 'LAST_NAME'])
                                    if fn or ln:
                                        first_name = first_name or fn or ''
                                        last_name = last_name or ln or ''
                                # If only last_name was provided (e.g. CSV had only "Last Name"), use it as first_name so name column shows the name
                                if last_name and not first_name:
                                    first_name = last_name
                                    last_name = ''
                                
                                # Handle phone - support both 'phone' and 'Mobile No'
                                phone_raw = get_value(['phone', 'Phone', 'PHONE', 'Mobile No', 'mobile no', 'MOBILE NO', 'Mobile'])
                                phone = ''
                                if phone_raw:
                                    # Handle scientific notation (e.g., 9.87654E+11)
                                    try:
                                        if 'E+' in str(phone_raw).upper() or 'e+' in str(phone_raw):
                                            phone = str(int(float(phone_raw)))
                                        else:
                                            phone = str(phone_raw).strip()
                                        
                                        # Remove all non-digit characters (keep only digits)
                                        phone = re.sub(r'\D', '', phone)
                                        
                                        # Handle Indian phone numbers (10 digits)
                                        # If it's 10 digits, keep as-is (will be normalized by serializer to +91XXXXXXXXXX)
                                        if len(phone) == 10:
                                            # Keep the 10-digit number as-is
                                            pass
                                        # If it starts with 91 and has 12 digits, remove the 91 prefix
                                        elif phone.startswith('91') and len(phone) == 12:
                                            phone = phone[2:]  # Remove country code, keep 10 digits
                                        # If it has 11 digits and starts with 0, remove the leading 0
                                        elif len(phone) == 11 and phone.startswith('0'):
                                            phone = phone[1:]  # Remove leading 0, keep 10 digits
                                        # If it's longer than 10 digits, take last 10
                                        elif len(phone) > 10:
                                            phone = phone[-10:]  # Take last 10 digits
                                        # If it's less than 10 digits, keep as-is (might be incomplete)
                                        
                                    except (ValueError, TypeError):
                                        phone = str(phone_raw).strip()
                                        # Clean to digits only
                                        phone = re.sub(r'\D', '', phone)
                                
                                # Get email (optional in new format)
                                email = get_value(['email', 'Email', 'EMAIL'])
                                # Clean email - if empty string, set to None
                                if email and not email.strip():
                                    email = None
                                
                                # Require either email or phone
                                if not email and not phone:
                                    msg = 'Either email or phone is required'
                                    counters['errors'].append(f'Row {row_num}: {msg}')
                                    row_error = msg
                                    yield {
                                        'row_num': row_num,
                                        'imported': counters['imported_count'],
                                        'skipped': counters['skipped_count'],
                                        'failed': len(counters['errors']),
                                        'name': f'{first_name} {last_name}'.strip(),
                                        'error': msg,
                                    }
                                    continue
                                
                                # Normalize phone number for imports - default to India (+91)
                                if phone:
                                    from shared.validators import normalize_phone_number
                                    digits_only = re.sub(r'\D', '', str(phone))
                                    
                                    # For imports, default to India (+91) if no country code
                                    if not str(phone).strip().startswith('+'):
                                        # If it's a 10-digit number, assume it's Indian
                                        if len(digits_only) == 10:
                                            phone = f'+91{digits_only}'
                                        # If it starts with 91 and has 12 digits, add + prefix
                                        elif digits_only.startswith('91') and len(digits_only) == 12:
                                            phone = f'+{digits_only}'
                                        # If it has 11 digits and starts with 0, remove 0 and add +91
                                        elif len(digits_only) == 11 and digits_only.startswith('0'):
                                            phone = f'+91{digits_only[1:]}'
                                        # For any other number without country code, assume Indian
                                        elif len(digits_only) >= 7 and len(digits_only) <= 12:
                                            phone = f'+91{digits_only}'
                                        else:
                                            # Use standard normalization as fallback
                                            phone = normalize_phone_number(phone)
                                    else:
                                        # Already has country code, normalize it
                                        phone = normalize_phone_number(phone)
                                
                                # Customer already exists (in DB or in-file): add this row as a visit, never skip.
                                existing_client_id = None
                                if email and email in existing_emails:
                                    existing_client_id = email_to_client_id.get(email)
                                if existing_client_id is None and phone and phone in existing_phones:
                                    existing_client_id = phone_to_client_id.get(phone)
                                if existing_client_id is not None or (email and email in existing_emails) or (phone and phone in existing_phones):
                                    store_id_v, visit_date_v = parse_visit_from_row(row, get_value)
                                    if existing_client_id and visit_date_v:
                                        try:
                                            # Support adding visits, but don't override: skip if this (client, store, date) already exists
                                            if not ClientVisit.objects.filter(
                                                client_id=existing_client_id,
                                                store_id=store_id_v,
                                                visit_date=visit_date_v,
                                            ).exists():
                                                ClientVisit.objects.create(
                                                    client_id=existing_client_id,
                                                    store_id=store_id_v,
                                                    visit_date=visit_date_v,
                                                    attended_by=get_value(['Attended By', 'attended_by', 'ATTENDED_BY', 'Sales Person']) or None,
                                                )
                                                counters['visits_added_count'] += 1
                                        except Exception as ev:
                                            counters['errors'].append(f'Row {row_num}: Could not add visit: {ev}')
                                            row_error = str(ev)
                                    elif not existing_client_id and visit_date_v:
                                        deferred_visits.append((email or None, phone or None, store_id_v, visit_date_v))
                                    yield {
                                        'row_num': row_num,
                                        'imported': counters['imported_count'],
                                        'skipped': counters['skipped_count'],
                                        'failed': len(counters['errors']),
                                        'name': f'{first_name} {last_name}'.strip(),
                                        'error': row_error,
                                    }
                                    continue
                                
                                # OPTIMIZED: Handle Store/Branch using pre-fetched stores
                                store_id = None
                                branch_name = get_value(['Branch', 'branch', 'BRANCH', 'Store', 'store', 'STORE'])
                                if branch_name and str(branch_name).strip():
                                    bn = str(branch_name).strip().lower()
                                    # Try exact match first
                                    store = tenant_stores.get(bn)
                                    if not store:
                                        # Try normalized match
                                        bn_normalized = re.sub(r'[^a-z0-9]', '', bn)
                                        store = normalized_stores.get(bn_normalized)
                                    if store:
                                        store_id = store.id
                                    else:
                                        store_id = import_store.id if import_store else None
                                else:
                                    store_id = import_store.id if import_store else None
                                
                                # Handle Status - map Close/Open to status values
                                status_raw = get_value(['Status', 'status', 'STATUS'], 'general')
                                status_value = 'general'
                                if status_raw.lower() in ['close', 'closed', 'closed_won']:
                                    status_value = 'vvip'  # Closed deals = VVIP
                                elif status_raw.lower() in ['open', 'active']:
                                    status_value = 'general'  # Open deals = General
                                elif status_raw.lower() in ['vvip', 'vip', 'general']:
                                    status_value = status_raw.lower()
                                
                                # Handle Preferred flag
                                preferred_raw = get_value(['Preferred', 'preferred', 'PREFERRED'], '')
                                preferred_flag = False
                                if preferred_raw.lower() in ['yes', 'y', 'true', '1']:
                                    preferred_flag = True
                                    if status_value == 'general':
                                        status_value = 'vip'  # Preferred customers default to VIP
                                
                                # OPTIMIZED: Handle assigned_to using pre-fetched users
                                assigned_to_value = None
                                assigned_to_user_obj = None
                                assigned_to_username = get_value(['assigned_to', 'Assigned To', 'ASSIGNED_TO', 'assigned_to', 'Attended By', 'attended_by', 'Sales Person'])
                                if assigned_to_username and assigned_to_username.strip():
                                    assigned_to_user_obj = tenant_users.get(assigned_to_username.strip().lower())
                                    if assigned_to_user_obj:
                                        assigned_to_value = assigned_to_user_obj
                                    else:
                                        if salesperson_not_found == 'skip':
                                            msg = f'Skipped (salesperson "{assigned_to_username}" not found)'
                                            counters['errors'].append(f'Row {row_num}: {msg}')
                                            row_error = msg
                                            yield {
                                                'row_num': row_num,
                                                'imported': counters['imported_count'],
                                                'skipped': counters['skipped_count'],
                                                'failed': len(counters['errors']),
                                                'name': f'{first_name} {last_name}'.strip(),
                                                'error': msg,
                                            }
                                            continue
                                        if salesperson_not_found == 'name_only':
                                            assigned_to_value = None
                                            assigned_to_user_obj = None
                                            # attended_by is set below from get_value
                                        elif salesperson_not_found == 'auto_create':
                                            try:
                                                new_username = assigned_to_username.strip()
                                                existing_user = tenant_users.get(new_username.lower())
                                                if existing_user:
                                                    assigned_to_value = existing_user
                                                    assigned_to_user_obj = existing_user
                                                else:
                                                    try:
                                                        new_user = User(
                                                            username=new_username,
                                                            first_name=assigned_to_username.strip(),
                                                            role='inhouse_sales',
                                                            tenant=request.user.tenant,
                                                            store=import_store,
                                                            is_active=True,
                                                        )
                                                        new_user.set_unusable_password()
                                                        new_user.save()
                                                        # Update cache
                                                        tenant_users[new_username.lower()] = new_user
                                                        assigned_to_value = new_user
                                                        assigned_to_user_obj = new_user
                                                    except IntegrityError:
                                                        existing_user = User.objects.filter(username=new_username, tenant=request.user.tenant).first()
                                                        if existing_user:
                                                            tenant_users[new_username.lower()] = existing_user
                                                            assigned_to_value = existing_user
                                                            assigned_to_user_obj = existing_user
                                                        else:
                                                            msg = f'Salesperson "{assigned_to_username}" already exists in another tenant; could not assign.'
                                                            counters['errors'].append(f'Row {row_num}: {msg}')
                                                            row_error = msg
                                                            yield {
                                                                'row_num': row_num,
                                                                'imported': counters['imported_count'],
                                                                'skipped': counters['skipped_count'],
                                                                'failed': len(counters['errors']),
                                                                'name': f'{first_name} {last_name}'.strip(),
                                                                'error': msg,
                                                            }
                                                            continue
                                            except Exception as e:
                                                msg = f'Could not auto-create salesperson "{assigned_to_username}": {e}'
                                                counters['errors'].append(f'Row {row_num}: {msg}')
                                                row_error = msg
                                                yield {
                                                    'row_num': row_num,
                                                    'imported': counters['imported_count'],
                                                    'skipped': counters['skipped_count'],
                                                    'failed': len(counters['errors']),
                                                    'name': f'{first_name} {last_name}'.strip(),
                                                    'error': msg,
                                                }
                                                continue
                                
                                # Prepare data for creation
                                client_data = {
                                    'first_name': first_name,
                                    'last_name': last_name,
                                    'email': email,  # Can be None if not provided
                                    'phone': phone or '',
                                    'customer_type': get_value(['customer_type', 'Customer Type'], 'individual'),
                                    'address': get_value(['address', 'Address', 'ADDRESS'], ''),
                                    'city': get_value(['City', 'city', 'CITY'], ''),
                                    'state': get_value(['State', 'state', 'STATE'], ''),
                                    'country': get_value(['Country', 'country', 'COUNTRY'], ''),
                                    'postal_code': get_value(['postal_code', 'Postal Code', 'postal_code', 'pincode', 'Pincode'], ''),
                                    'preferred_metal': get_value(['preferred_metal', 'Preferred Metal', 'item_category', 'Item Category'], ''),
                                    'preferred_stone': get_value(['preferred_stone', 'Preferred Stone'], ''),
                                    'ring_size': get_value(['ring_size', 'Ring Size'], ''),
                                    'budget_range': get_value(['budget_range', 'Budget Range'], ''),
                                    'lead_source': get_value(['lead_source', 'Lead Source'], ''),
                                    'notes': get_value(['notes', 'Notes', 'NOTES'], ''),
                                    'community': get_value(['community', 'Community'], ''),
                                    'mother_tongue': get_value(['mother_tongue', 'Mother Tongue'], ''),
                                    'reason_for_visit': get_value(['reason_for_visit', 'Reason for Visit'], ''),
                                    'age_of_end_user': get_value(['age_of_end_user', 'Age of End User'], ''),
                                    'saving_scheme': get_value(['saving_scheme', 'Saving Scheme'], ''),
                                    'catchment_area': get_value(['catchment_area', 'Catchment Area', 'Area', 'area', 'AREA'], ''),
                                    'next_follow_up': get_value(['next_follow_up', 'Next Follow Up'], ''),
                                    'summary_notes': get_value(['summary_notes', 'Summary Notes'], ''),
                                    'status': status_value,
                                    'tenant': request.user.tenant.id if request.user.tenant else None,
                                    'store': store_id,
                                    # New fields from CSV
                                    # 'sr_no': get_value(['SR.NO', 'SR_NO', 'sr_no', 'SR No', 'Reference ID'], ''),  # Temporarily commented - uncomment after running migration 0031
                                    'area': get_value(['Area', 'area', 'AREA'], ''),
                                    'client_category': get_value(['Client Category', 'client_category', 'CLIENT_CATEGORY'], ''),
                                    'preferred_flag': preferred_flag,
                                    'attended_by': get_value(['Attended By', 'attended_by', 'ATTENDED_BY', 'Sales Person'], ''),
                                    'item_category': get_value(['Item Category', 'item_category', 'ITEM_CATEGORY'], ''),
                                    'item_name': get_value(['Item Name', 'item_name', 'ITEM_NAME'], ''),
                                }
                                # Set assigned_to as User object (not username string) for direct creation
                                assigned_to_user = assigned_to_user_obj if assigned_to_user_obj else None
                                
                                if salesperson_not_found == 'name_only' and assigned_to_username and not assigned_to_user:
                                    client_data['attended_by'] = assigned_to_username.strip()
                                if client_data.get('attended_by') and not client_data.get('sales_person'):
                                    client_data['sales_person'] = client_data['attended_by']
                                
                                # Set default values for import to avoid validation errors
                                # These are optional but we provide defaults for better data quality
                                if not client_data.get('state') or not str(client_data.get('state', '')).strip():
                                    # Use city as state if available, otherwise leave empty (will be None)
                                    city_val = client_data.get('city', '').strip()
                                    client_data['state'] = city_val if city_val else None
                                else:
                                    # Clean state value
                                    state_val = str(client_data.get('state', '')).strip()
                                    client_data['state'] = state_val if state_val else None
                                
                                if not client_data.get('catchment_area') or not str(client_data.get('catchment_area', '')).strip():
                                    # Use 'area' field if available, otherwise use city
                                    area_val = client_data.get('area', '').strip()
                                    city_val = client_data.get('city', '').strip()
                                    client_data['catchment_area'] = area_val or city_val or None
                                else:
                                    # Clean catchment_area value
                                    catchment_val = str(client_data.get('catchment_area', '')).strip()
                                    client_data['catchment_area'] = catchment_val if catchment_val else None
                                
                                if not client_data.get('lead_source') or not str(client_data.get('lead_source', '')).strip():
                                    # Use client_category or default
                                    category_val = client_data.get('client_category', '').strip()
                                    client_data['lead_source'] = category_val if category_val else None
                                else:
                                    # Clean lead_source value
                                    lead_val = str(client_data.get('lead_source', '')).strip()
                                    client_data['lead_source'] = lead_val if lead_val else None
                                
                                if not client_data.get('reason_for_visit') or not str(client_data.get('reason_for_visit', '')).strip():
                                    # Default reason for imported customers
                                    client_data['reason_for_visit'] = None  # Leave empty for imports
                                else:
                                    # Clean reason_for_visit value
                                    reason_val = str(client_data.get('reason_for_visit', '')).strip()
                                    client_data['reason_for_visit'] = reason_val if reason_val else None
                                
                                if not client_data.get('product_type') or not str(client_data.get('product_type', '')).strip():
                                    # Use item_category or item_name if available
                                    item_cat = client_data.get('item_category', '').strip()
                                    item_name = client_data.get('item_name', '').strip()
                                    client_data['product_type'] = item_cat or item_name or None
                                else:
                                    # Clean product_type value
                                    product_val = str(client_data.get('product_type', '')).strip()
                                    client_data['product_type'] = product_val if product_val else None
                                
                                # Clean all string fields - convert empty strings to None
                                string_fields = [
                                    'address', 'country', 'postal_code', 'preferred_metal', 'preferred_stone',
                                    'ring_size', 'budget_range', 'notes', 'community', 'mother_tongue',
                                    'age_of_end_user', 'saving_scheme', 'summary_notes', 'sales_person',
                                    'customer_status', 'style', 'material_type', 'product_subtype',
                                    'gold_range', 'diamond_range', 'customer_preferences', 'design_selected',
                                    'wants_more_discount', 'checking_other_jewellers', 'let_him_visit', 'design_number',
                                    # 'sr_no',  # Temporarily commented - uncomment after running migration 0031
                                    'area', 'client_category', 'attended_by', 'item_category', 'item_name'
                                ]
                                for field in string_fields:
                                    if field in client_data:
                                        val = client_data[field]
                                        if isinstance(val, str) and not val.strip():
                                            client_data[field] = None
                                        elif val == '':
                                            client_data[field] = None
                                
                                # Handle date fields - support multiple formats
                                def parse_date(date_str, formats=['%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y']):
                                    """Parse date string with multiple format support"""
                                    if not date_str or date_str.strip() == '':
                                        return None
                                    date_str = date_str.strip()
                                    for fmt in formats:
                                        try:
                                            return datetime.strptime(date_str, fmt).date()
                                        except ValueError:
                                            continue
                                    return None
                                
                                # Handle date_of_birth
                                date_of_birth = get_value(['date_of_birth', 'Date of Birth', 'DOB', 'dob'])
                                if date_of_birth:
                                    parsed_dob = parse_date(date_of_birth)
                                    if parsed_dob:
                                        client_data['date_of_birth'] = parsed_dob
                                
                                # Handle anniversary_date
                                anniversary_date = get_value(['anniversary_date', 'Anniversary Date', 'Anniversary'])
                                if anniversary_date:
                                    parsed_anniv = parse_date(anniversary_date)
                                    if parsed_anniv:
                                        client_data['anniversary_date'] = parsed_anniv
                                
                                # Handle visit_date (new field from CSV)
                                visit_date = get_value(['Visit Date', 'visit_date', 'VISIT_DATE', 'Visit'])
                                if visit_date:
                                    parsed_visit = parse_date(visit_date)
                                    if parsed_visit:
                                        client_data['visit_date'] = parsed_visit
                                
                                # Handle next_follow_up
                                next_follow_up = get_value(['next_follow_up', 'Next Follow Up', 'Next Follow-Up'])
                                if next_follow_up:
                                    parsed_followup = parse_date(next_follow_up)
                                    if parsed_followup:
                                        client_data['next_follow_up'] = parsed_followup
                                
                                # created_at from CSV: REQUIRED - use only the value from the column, NEVER default to today
                                # Check if created_at column exists in CSV (check multiple possible column names)
                                created_at_str = None
                                created_at_column_found = False
                                for col_name in ['created_at', 'Created At', 'Created_at', 'CREATED_AT', 'Date', 'date', 'created_date', 'Created Date']:
                                    if col_name in row or any(rk.lower().strip() == col_name.lower().strip() for rk in row.keys()):
                                        created_at_column_found = True
                                        val = get_value([col_name])
                                        if val and str(val).strip():
                                            created_at_str = str(val).strip()
                                            break
                                
                                # If column exists but is empty, or column doesn't exist at all, add error
                                if not created_at_column_found:
                                    msg = 'created_at column is missing from CSV. This column is required for import.'
                                    counters['errors'].append(f'Row {row_num}: {msg}')
                                    row_error = msg
                                    yield {
                                        'row_num': row_num,
                                        'imported': counters['imported_count'],
                                        'skipped': counters['skipped_count'],
                                        'failed': len(counters['errors']),
                                        'name': f'{first_name} {last_name}'.strip(),
                                        'error': msg,
                                    }
                                    continue
                                
                                if not created_at_str:
                                    msg = 'created_at column exists but is empty. A valid date is required for import.'
                                    counters['errors'].append(f'Row {row_num}: {msg}')
                                    row_error = msg
                                    yield {
                                        'row_num': row_num,
                                        'imported': counters['imported_count'],
                                        'skipped': counters['skipped_count'],
                                        'failed': len(counters['errors']),
                                        'name': f'{first_name} {last_name}'.strip(),
                                        'error': msg,
                                    }
                                    continue
                                
                                # Parse created_at (must be datetime, not date)
                                created_at_dt = None
                                try:
                                    # Try parsing as datetime first (with time)
                                    created_at_dt = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                                    created_at_dt = timezone.make_aware(created_at_dt)
                                except:
                                    try:
                                        # Try parsing as datetime with different format
                                        created_at_dt = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M')
                                        created_at_dt = timezone.make_aware(created_at_dt)
                                    except:
                                        try:
                                            # Try parsing as date (YYYY-MM-DD) and convert to datetime at midnight
                                            date_obj = datetime.strptime(created_at_str, '%Y-%m-%d').date()
                                            created_at_dt = timezone.make_aware(datetime.combine(date_obj, datetime.min.time()))
                                        except:
                                            try:
                                                # Try DD-MM-YYYY format
                                                date_obj = datetime.strptime(created_at_str, '%d-%m-%Y').date()
                                                created_at_dt = timezone.make_aware(datetime.combine(date_obj, datetime.min.time()))
                                            except:
                                                try:
                                                    # Try DD/MM/YYYY format
                                                    date_obj = datetime.strptime(created_at_str, '%d/%m/%Y').date()
                                                    created_at_dt = timezone.make_aware(datetime.combine(date_obj, datetime.min.time()))
                                                except:
                                                    msg = f'created_at value "{created_at_str}" could not be parsed as a valid date. Expected formats: YYYY-MM-DD, DD-MM-YYYY, or DD/MM/YYYY.'
                                                    counters['errors'].append(f'Row {row_num}: {msg}')
                                                    row_error = msg
                                                    yield {
                                                        'row_num': row_num,
                                                        'imported': counters['imported_count'],
                                                        'skipped': counters['skipped_count'],
                                                        'failed': len(counters['errors']),
                                                        'name': f'{first_name} {last_name}'.strip(),
                                                        'error': msg,
                                                    }
                                                    continue
                                
                                # Clean data - remove empty strings and None values
                                cleaned_data = {}
                                for key, value in client_data.items():
                                    if value is not None and value != '':
                                        cleaned_data[key] = value
                                    elif value == '':
                                        # Skip empty strings
                                        pass
                                
                                # OPTIMIZED: Create Client object directly instead of using serializer (much faster)
                                try:
                                    # Create Client object directly (skip serializer overhead)
                                    # Note: bulk_create doesn't call save() so auto_now_add won't work - set manually
                                    # created_at_dt is already parsed and validated above
                                    now = timezone.now()
                                    client = Client(
                                        first_name=cleaned_data.get('first_name') or '',
                                        last_name=cleaned_data.get('last_name') or '',
                                        email=cleaned_data.get('email'),
                                        phone=cleaned_data.get('phone') or '',
                                        assigned_to=assigned_to_user,
                                        created_by=request.user,
                                        tenant=request.user.tenant,
                                        store_id=store_id or (import_store.id if import_store else None),
                                        customer_type=cleaned_data.get('customer_type', 'individual'),
                                        address=cleaned_data.get('address'),
                                        city=cleaned_data.get('city'),
                                        state=cleaned_data.get('state'),
                                        country=cleaned_data.get('country'),
                                        postal_code=cleaned_data.get('postal_code'),
                                        preferred_metal=cleaned_data.get('preferred_metal'),
                                        preferred_stone=cleaned_data.get('preferred_stone'),
                                        ring_size=cleaned_data.get('ring_size'),
                                        budget_range=cleaned_data.get('budget_range'),
                                        lead_source=cleaned_data.get('lead_source'),
                                        notes=cleaned_data.get('notes'),
                                        community=cleaned_data.get('community'),
                                        mother_tongue=cleaned_data.get('mother_tongue'),
                                        reason_for_visit=cleaned_data.get('reason_for_visit'),
                                        age_of_end_user=cleaned_data.get('age_of_end_user'),
                                        saving_scheme=cleaned_data.get('saving_scheme'),
                                        catchment_area=cleaned_data.get('catchment_area'),
                                        next_follow_up=cleaned_data.get('next_follow_up'),
                                        summary_notes=cleaned_data.get('summary_notes'),
                                        status=cleaned_data.get('status', 'general'),
                                        date_of_birth=cleaned_data.get('date_of_birth'),
                                        anniversary_date=cleaned_data.get('anniversary_date'),
                                        pincode=cleaned_data.get('pincode'),
                                        sales_person=cleaned_data.get('sales_person') or cleaned_data.get('attended_by'),
                                        product_type=cleaned_data.get('product_type'),
                                        area=cleaned_data.get('area'),
                                        client_category=cleaned_data.get('client_category'),
                                        preferred_flag=cleaned_data.get('preferred_flag', False),
                                        attended_by=cleaned_data.get('attended_by'),
                                        item_category=cleaned_data.get('item_category'),
                                        item_name=cleaned_data.get('item_name'),
                                        created_at=created_at_dt,  # Use ONLY the date from CSV (already validated above)
                                        updated_at=now,  # bulk_create doesn't set auto_now
                                    )
                                    
                                    # Add to batch for bulk_create
                                    clients_to_create.append(client)
                                    
                                    # Update existing sets to prevent duplicates in same batch
                                    if email:
                                        existing_emails.add(email)
                                    if phone:
                                        existing_phones.add(phone)
                                    
                                    # Bulk create when batch is full
                                    if len(clients_to_create) >= BATCH_SIZE:
                                        # CRITICAL: Store CSV created_at values BEFORE bulk_create
                                        # because bulk_create() might overwrite them with today's date
                                        csv_created_at_map = {}
                                        for idx, client in enumerate(clients_to_create):
                                            if client.created_at:
                                                csv_created_at_map[idx] = client.created_at
                                        
                                        Client.objects.bulk_create(clients_to_create, ignore_conflicts=False)
                                        
                                        # Update created_at using queryset.update() which DEFINITELY bypasses auto_now_add
                                        # Group by created_at value to minimize queries (same date = one update query)
                                        created_at_groups = {}
                                        for idx, client in enumerate(clients_to_create):
                                            if client.pk and idx in csv_created_at_map:
                                                # Use the ORIGINAL CSV value, not what Django might have set
                                                csv_created_at = csv_created_at_map[idx]
                                                if csv_created_at not in created_at_groups:
                                                    created_at_groups[csv_created_at] = []
                                                created_at_groups[csv_created_at].append(client.pk)
                                        
                                        # Update each group in a single query
                                        for created_at_dt, pk_list in created_at_groups.items():
                                            Client.objects.filter(pk__in=pk_list).update(created_at=created_at_dt)
                                        # Record one ClientVisit per new client (same store/date as row)
                                        for idx, c in enumerate(clients_to_create):
                                            if c.pk:
                                                visit_date_val = csv_created_at_map.get(idx)
                                                if visit_date_val:
                                                    vd = visit_date_val.date() if hasattr(visit_date_val, 'date') else visit_date_val
                                                    try:
                                                        ClientVisit.objects.create(
                                                            client_id=c.pk,
                                                            store_id=c.store_id,
                                                            visit_date=vd,
                                                            attended_by=getattr(c, 'attended_by', None),
                                                        )
                                                    except Exception:
                                                        pass
                                                if c.email:
                                                    email_to_client_id[c.email] = c.pk
                                                    existing_emails.add(c.email)
                                                if c.phone:
                                                    phone_to_client_id[c.phone] = c.pk
                                                    existing_phones.add(c.phone)
                                        counters['imported_count'] += len(clients_to_create)
                                        clients_to_create = []
                                    
                                except Exception as save_error:
                                    error_detail = str(save_error)
                                    msg = f'Failed to create customer: {error_detail}'
                                    counters['errors'].append(f'Row {row_num}: {msg}')
                                    row_error = msg
                            
                            except Exception as e:
                                err_msg = str(e) if e else 'Unknown error'
                                if hasattr(e, 'detail'):
                                    err_msg = str(e.detail)
                                counters['errors'].append(f'Row {row_num}: {err_msg}')
                                row_error = err_msg

                            # Yield per-row status for streaming callers
                            yield {
                                'row_num': row_num,
                                'imported': counters['imported_count'] + len(clients_to_create),  # Include pending batch
                                'skipped': counters['skipped_count'],
                                'failed': len(counters['errors']),
                                'name': f'{first_name} {last_name}'.strip(),
                                'error': row_error,
                            }
                    
                    # Final bulk_create for remaining clients
                    if clients_to_create:
                        # CRITICAL: Store CSV created_at values BEFORE bulk_create
                        # because bulk_create() might overwrite them with today's date
                        csv_created_at_map = {}
                        for idx, client in enumerate(clients_to_create):
                            if client.created_at:
                                csv_created_at_map[idx] = client.created_at
                        
                        Client.objects.bulk_create(clients_to_create, ignore_conflicts=False)
                        
                        # Update created_at using queryset.update() which DEFINITELY bypasses auto_now_add
                        # Group by created_at value to minimize queries (same date = one update query)
                        created_at_groups = {}
                        for idx, client in enumerate(clients_to_create):
                            if client.pk and idx in csv_created_at_map:
                                # Use the ORIGINAL CSV value, not what Django might have set
                                csv_created_at = csv_created_at_map[idx]
                                if csv_created_at not in created_at_groups:
                                    created_at_groups[csv_created_at] = []
                                created_at_groups[csv_created_at].append(client.pk)
                        
                        # Update each group in a single query
                        for created_at_dt, pk_list in created_at_groups.items():
                            Client.objects.filter(pk__in=pk_list).update(created_at=created_at_dt)
                        # Record one ClientVisit per new client and update maps
                        for idx, c in enumerate(clients_to_create):
                            if c.pk:
                                visit_date_val = csv_created_at_map.get(idx)
                                if visit_date_val:
                                    vd = visit_date_val.date() if hasattr(visit_date_val, 'date') else visit_date_val
                                    try:
                                        ClientVisit.objects.create(
                                            client_id=c.pk,
                                            store_id=c.store_id,
                                            visit_date=vd,
                                            attended_by=getattr(c, 'attended_by', None),
                                        )
                                    except Exception:
                                        pass
                                if c.email:
                                    email_to_client_id[c.email] = c.pk
                                    existing_emails.add(c.email)
                                if c.phone:
                                    phone_to_client_id[c.phone] = c.pk
                                    existing_phones.add(c.phone)
                        counters['imported_count'] += len(clients_to_create)
                        clients_to_create = []

                    # Process deferred visits (in-file duplicates: we now have client_id from flushed batches)
                    for def_email, def_phone, def_store_id, def_visit_date in deferred_visits:
                        cid = None
                        if def_email:
                            cid = email_to_client_id.get(def_email)
                        if cid is None and def_phone:
                            cid = phone_to_client_id.get(def_phone)
                        if cid and def_visit_date:
                            try:
                                # Support adding visits, but don't override: skip if this (client, store, date) already exists
                                if not ClientVisit.objects.filter(
                                    client_id=cid,
                                    store_id=def_store_id,
                                    visit_date=def_visit_date,
                                ).exists():
                                    ClientVisit.objects.create(
                                        client_id=cid,
                                        store_id=def_store_id,
                                        visit_date=def_visit_date,
                                    )
                                    counters['visits_added_count'] += 1
                            except Exception:
                                pass

                return counters, row_iterator()

            stream = request.GET.get('stream') == '1' or request.query_params.get('stream') == '1'
            total_rows = len(rows)

            # Streaming mode using text/event-stream so the frontend can show true progress
            if stream:
                counters, row_gen = perform_import(stream_progress=True)

                def generate():
                    # Initial event
                    yield f"data: {json.dumps({'type': 'progress', 'processed': 0, 'total': total_rows, 'imported': 0, 'skipped': 0, 'failed': 0})}\n\n"
                    for row_result in row_gen:
                        processed = row_result.get('row_num', 0) - 1
                        payload = {
                            'type': 'progress',
                            'processed': processed,
                            'total': total_rows,
                            'imported': row_result.get('imported', 0),
                            'skipped': row_result.get('skipped', 0),
                            'failed': row_result.get('failed', 0),
                            'row_num': row_result.get('row_num'),
                            'name': row_result.get('name') or '',
                            'error': row_result.get('error'),
                        }
                        yield f"data: {json.dumps(payload)}\n\n"

                    # After generator completes, write audit and final "done" event
                    try:
                        details = {
                            'salesperson_not_found': salesperson_not_found,
                            'errors_sample': counters['errors'][:20],
                            'errors': counters['errors'][:MAX_AUDIT_ERRORS_STORED],
                        }
                        CustomerImportAudit.objects.create(
                            user=request.user,
                            action='imported',
                            total_rows=total_rows,
                            valid_count=counters['imported_count'] + len(counters['errors']),
                            invalid_count=0,
                            needs_attention_count=0,
                            imported_count=counters['imported_count'],
                            failed_count=len(counters['errors']),
                            details=details,
                        )
                    except Exception as audit_err:
                        logger.warning('Could not write CustomerImportAudit (run migrations?): %s', audit_err)

                    done_payload = {
                        'type': 'done',
                        'total_rows': total_rows,
                        'imported': counters['imported_count'],
                        'skipped': counters['skipped_count'],
                        'visits_added': counters.get('visits_added_count', 0),
                        'failed': len(counters['errors']),
                        'errors': counters['errors'],
                    }
                    yield f"data: {json.dumps(done_payload)}\n\n"

                return StreamingHttpResponse(
                    generate(),
                    content_type='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
                )

            # Non-streaming (original) behaviour
            counters, row_gen = perform_import(stream_progress=False)
            # Exhaust generator to perform the import side-effects
            for _ in row_gen:
                pass

            try:
                details = {
                    'salesperson_not_found': salesperson_not_found,
                    'errors_sample': counters['errors'][:20],
                    'errors': counters['errors'][:MAX_AUDIT_ERRORS_STORED],
                }
                CustomerImportAudit.objects.create(
                    user=request.user,
                    action='imported',
                    total_rows=len(rows),
                    valid_count=counters['imported_count'] + len(counters['errors']),
                    invalid_count=0,
                    needs_attention_count=0,
                    imported_count=counters['imported_count'],
                    failed_count=len(counters['errors']),
                    details=details,
                )
            except Exception as audit_err:
                logger.warning('Could not write CustomerImportAudit (run migrations?): %s', audit_err)
            return Response({
                'message': f'Import completed. {counters["imported_count"]} customers imported successfully.',
                'imported': counters['imported_count'],
                'skipped': counters['skipped_count'],
                'visits_added': counters.get('visits_added_count', 0),
                'failed': len(counters['errors']),
                'errors': counters['errors']
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.exception('import_file failed')
            return Response(
                {'error': f'Import failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], permission_classes=[ImportExportPermission])
    def import_audits(self, request):
        """Return recent customer import/validation audits for the current user."""
        limit = min(int(request.query_params.get('limit', 10)), 50)
        audits = CustomerImportAudit.objects.filter(user=request.user).order_by('-created_at')[:limit]
        return Response({
            'results': [
                {
                    'id': a.id,
                    'action': a.action,
                    'total_rows': a.total_rows,
                    'valid_count': a.valid_count,
                    'invalid_count': a.invalid_count,
                    'needs_attention_count': a.needs_attention_count,
                    'imported_count': a.imported_count,
                    'failed_count': a.failed_count,
                    'created_at': a.created_at.isoformat() if a.created_at else None,
                }
                for a in audits
            ],
            'max_batch_size': MAX_IMPORT_BATCH_SIZE,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], permission_classes=[ImportExportPermission])
    def import_json(self, request):
        """Import customers from JSON - only for business admin and managers"""
        try:
            if 'file' not in request.FILES:
                return Response(
                    {'error': 'No file provided'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            json_file = request.FILES['file']
            
            # Validate file type
            if not json_file.name.endswith('.json'):
                return Response(
                    {'error': 'Please upload a JSON file'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Read JSON file
            json_data = json.loads(json_file.read().decode('utf-8'))
            
            if not isinstance(json_data, list):
                return Response(
                    {'error': 'JSON file should contain an array of customer objects'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            imported_count = 0
            errors = []
            
            with transaction.atomic():
                for row_num, customer_data in enumerate(json_data, start=1):
                    try:
                        # Clean and validate data - email is now optional
                        email = customer_data.get('email', '').strip()
                        phone = customer_data.get('phone', '').strip()
                        
                        # Require either email or phone for identification
                        if not email and not phone:
                            errors.append(f'Row {row_num}: Either email or phone is required')
                            continue
                        
                        # Check if customer already exists (by email if provided)
                        if email and Client.objects.filter(email=email, tenant=request.user.tenant, is_deleted=False).exists():
                            errors.append(f'Row {row_num}: Customer with email {email} already exists')
                            continue
                        
                        # Prepare data for creation
                        client_data = {
                            'first_name': customer_data.get('first_name', '').strip(),
                            'last_name': customer_data.get('last_name', '').strip(),
                            'email': email,
                            'phone': customer_data.get('phone', '').strip(),
                            'customer_type': customer_data.get('customer_type', 'individual'),
                            'address': customer_data.get('address', '').strip(),
                            'city': customer_data.get('city', '').strip(),
                            'state': customer_data.get('state', '').strip(),
                            'country': customer_data.get('country', '').strip(),
                            'postal_code': customer_data.get('postal_code', '').strip(),
                            'preferred_metal': customer_data.get('preferred_metal', '').strip(),
                            'preferred_stone': customer_data.get('preferred_stone', '').strip(),
                            'ring_size': customer_data.get('ring_size', '').strip(),
                            'budget_range': customer_data.get('budget_range', '').strip(),
                            'lead_source': customer_data.get('lead_source', '').strip(),
                            'notes': customer_data.get('notes', '').strip(),
                            'community': customer_data.get('community', '').strip(),
                            'mother_tongue': customer_data.get('mother_tongue', '').strip(),
                            'reason_for_visit': customer_data.get('reason_for_visit', '').strip(),
                            'age_of_end_user': customer_data.get('age_of_end_user', '').strip(),
                            'saving_scheme': customer_data.get('saving_scheme', '').strip(),
                            'catchment_area': customer_data.get('catchment_area', '').strip(),
                            'next_follow_up': customer_data.get('next_follow_up', '').strip(),
                            'summary_notes': customer_data.get('summary_notes', '').strip(),
                            'status': customer_data.get('status', 'general'),
                            'tenant': request.user.tenant.id if request.user.tenant else None,
                        }
                        
                        # Create client
                        serializer = self.get_serializer(data=client_data)
                        if serializer.is_valid():
                            client = serializer.save()
                            imported_count += 1
                        else:
                            errors.append(f'Row {row_num}: {serializer.errors}')
                    
                    except Exception as e:
                        errors.append(f'Row {row_num}: {str(e)}')
            
            return Response({
                'message': f'Import completed. {imported_count} customers imported successfully.',
                'imported_count': imported_count,
                'errors': errors
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Import failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], permission_classes=[ImportExportPermission])
    def download_template(self, request):
        """Download CSV or Excel template for import - only for business admin and managers"""
        try:
            # Get format preference from query parameter
            format_type = request.GET.get('format', 'csv').lower()
            
            if format_type == 'xlsx':
                # Create Excel template
                import openpyxl
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Customer Import Template"
                
                # Define headers
                headers = [
                    'first_name', 'last_name', 'email', 'phone', 'customer_type',
                    'address', 'city', 'state', 'country', 'postal_code',
                    'date_of_birth', 'anniversary_date', 'preferred_metal', 'preferred_stone',
                    'ring_size', 'budget_range', 'lead_source', 'notes', 'community',
                    'mother_tongue', 'reason_for_visit', 'age_of_end_user', 'saving_scheme',
                    'catchment_area', 'next_follow_up', 'summary_notes', 'status', 'tags'
                ]
                
                # Write headers
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Add example row
                example_data = [
                    'John', 'Doe', 'john.doe@example.com', '+1234567890', 'individual',
                    '123 Main St', 'New York', 'NY', 'USA', '10001',
                    '1990-01-01', '2015-06-15', 'Gold', 'Diamond',
                    '7', '1000-5000', 'website', 'Interested in engagement rings', 'General',
                    'English', 'Engagement Ring', '25-35', 'Monthly',
                    'Downtown', 'Call next week', 'High potential customer', 'general', 'High Value, Engagement, Gold'
                ]
                
                for col, value in enumerate(example_data, 1):
                    ws.cell(row=2, column=col, value=value)
                
                # Create response
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="customer_import_template.xlsx"'
                
                wb.save(response)
                return response
                
            else:
                # Default to CSV template
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="customer_import_template.csv"'
                
                # Define CSV fields
                fieldnames = [
                    'first_name', 'last_name', 'email', 'phone', 'customer_type',
                    'address', 'city', 'state', 'country', 'postal_code',
                    'date_of_birth', 'anniversary_date', 'preferred_metal', 'preferred_stone',
                    'ring_size', 'budget_range', 'lead_source', 'notes', 'community',
                    'mother_tongue', 'reason_for_visit', 'age_of_end_user', 'saving_scheme',
                    'catchment_area', 'next_follow_up', 'summary_notes', 'status', 'tags'
                ]
                
                writer = csv.DictWriter(response, fieldnames=fieldnames)
                writer.writeheader()
                
                # Add example row
                example_row = {
                    'first_name': 'John',
                    'last_name': 'Doe',
                    'email': 'john.doe@example.com',
                    'phone': '+1234567890',
                    'customer_type': 'individual',
                    'address': '123 Main St',
                    'city': 'New York',
                    'state': 'NY',
                    'country': 'USA',
                    'postal_code': '10001',
                    'date_of_birth': '1990-01-01',
                    'anniversary_date': '2015-06-15',
                    'preferred_metal': 'Gold',
                    'preferred_stone': 'Diamond',
                    'ring_size': '7',
                    'budget_range': '1000-5000',
                    'lead_source': 'website',
                    'notes': 'Interested in engagement rings',
                    'community': 'General',
                    'mother_tongue': 'English',
                    'reason_for_visit': 'Engagement Ring',
                    'age_of_end_user': '25-35',
                    'saving_scheme': 'Monthly',
                    'catchment_area': 'Downtown',
                    'next_follow_up': 'Call next week',
                    'summary_notes': 'High potential customer',
                    'status': 'general',
                    'tags': 'High Value, Engagement, Gold'
                }
                writer.writerow(example_row)
                
                return response
                
        except Exception as e:
            return Response(
                {'error': f'Template download failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='cross-store')
    def get_cross_store(self, request, pk=None):
        """
        Get customer by ID, bypassing store filtering for cross-store access within tenant.
        This is used when adding interests to existing customers from different stores.
        """
        try:
            # Get customer by ID within tenant (bypass store filtering)
            customer = Client.objects.filter(
                id=pk,
                tenant=request.user.tenant,
                is_deleted=False
            ).first()
            
            if not customer:
                return Response(
                    {'error': 'Customer not found or you do not have access'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Serialize the customer
            serializer = self.get_serializer(customer)
            return Response({
                'success': True,
                'data': serializer.data
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['put', 'patch'], url_path='cross-store', url_name='cross-store-update')
    def update_cross_store(self, request, pk=None):
        """
        Update customer by ID, bypassing store filtering for cross-store access within tenant.
        This is used when updating existing customers from different stores (e.g., adding interests).
        """
        try:
            # Get customer by ID within tenant (bypass store filtering)
            customer = Client.objects.filter(
                id=pk,
                tenant=request.user.tenant,
                is_deleted=False
            ).first()
            
            if not customer:
                return Response(
                    {'error': 'Customer not found or you do not have access'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Use the same serializer and update logic as regular update
            # For PUT, we need to handle partial updates
            partial = request.method == 'PATCH'
            serializer = self.get_serializer(customer, data=request.data, partial=partial)
            
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            print(f"Error updating customer cross-store: {e}")
            print(traceback.format_exc())
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='journey')
    def customer_journey(self, request, pk=None):
        """
        Get customer journey/history - timeline of all activities
        Returns: timeline of interests, interactions, appointments, pipeline entries, sales
        """
        try:
            # Use cross-store access to get customer (bypass store filtering)
            client = Client.objects.filter(
                id=pk,
                tenant=request.user.tenant,
                is_deleted=False
            ).first()
            
            if not client:
                return Response(
                    {'error': 'Customer not found or you do not have access'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            from apps.sales.models import SalesPipeline, Sale
            from django.utils import timezone
            from datetime import datetime
            import json
            import re
            
            journey_items = []
            
            # Helper function to extract design number from notes
            def extract_design_number(notes):
                if not notes:
                    return None
                match = re.search(r'Design Number:\s*([^\n.]+)', notes, re.IGNORECASE)
                return match.group(1).strip() if match else None
            
            # Helper function to extract images from notes
            def extract_images(notes):
                if not notes:
                    return []
                images = []
                match = re.search(r'Images:\s*(\[.*?\])', notes, re.IGNORECASE | re.DOTALL)
                if match:
                    try:
                        images = json.loads(match.group(1))
                    except:
                        pass
                return images
            
            # 1. Customer Interests (with store info from sales rep)
            interests = client.interests.all().select_related('category', 'product')
            for interest in interests:
                # Get store from sales rep who created the interest (if available)
                store_name = "Unknown Store"
                sales_rep_name = "Unknown"
                
                # Try multiple methods to find store and sales rep:
                # Method 1: Find pipeline created around the same time as interest
                # Prioritize pipelines created AFTER the interest (within 1 hour), then before (within 2 hours)
                related_pipeline = SalesPipeline.objects.filter(
                    client=client,
                    created_at__gte=interest.created_at,
                    created_at__lte=interest.created_at + timezone.timedelta(hours=1)
                ).order_by('created_at').first()  # Get the first one after interest (closest match)
                
                # If no pipeline found after interest, try before (within 2 hours)
                if not related_pipeline:
                    related_pipeline = SalesPipeline.objects.filter(
                        client=client,
                        created_at__gte=interest.created_at - timezone.timedelta(hours=2),
                        created_at__lt=interest.created_at
                    ).order_by('-created_at').first()  # Get the most recent one before interest
                
                # If still no pipeline found, try wider window (24 hours)
                if not related_pipeline:
                    related_pipeline = SalesPipeline.objects.filter(
                        client=client,
                        created_at__gte=interest.created_at - timezone.timedelta(hours=24),
                        created_at__lte=interest.created_at + timezone.timedelta(hours=24)
                    ).order_by('-created_at').first()
                
                if related_pipeline:
                    store_name = related_pipeline.sales_representative.store.name if related_pipeline.sales_representative.store else "No Store"
                    sales_rep_name = related_pipeline.sales_representative.get_full_name() or related_pipeline.sales_representative.username
                else:
                    # Method 2: Get store from client's store field (if available)
                    if client.store:
                        store_name = client.store.name
                    # Method 3: Try to find any pipeline for this client and use the most recent one's store
                    latest_pipeline = SalesPipeline.objects.filter(client=client).order_by('-created_at').first()
                    if latest_pipeline and latest_pipeline.sales_representative:
                        if not store_name or store_name == "Unknown Store":
                            store_name = latest_pipeline.sales_representative.store.name if latest_pipeline.sales_representative.store else "No Store"
                        sales_rep_name = latest_pipeline.sales_representative.get_full_name() or latest_pipeline.sales_representative.username
                
                journey_items.append({
                    'type': 'interest',
                    'id': interest.id,
                    'date': interest.created_at.isoformat() if interest.created_at else None,
                    'title': f"Product Interest Added",
                    'description': f"{interest.category.name if interest.category else 'Unknown'} - {interest.product.name if interest.product else 'Unknown'}",
                    'details': {
                        'category': interest.category.name if interest.category else None,
                        'product': interest.product.name if interest.product else None,
                        'revenue': float(interest.revenue) if interest.revenue else 0,
                        'store': store_name,
                        'sales_rep': sales_rep_name,
                        'design_number': extract_design_number(interest.notes),
                        'images': extract_images(interest.notes),
                        'is_purchased': interest.is_purchased,
                        'is_not_purchased': interest.is_not_purchased,
                    }
                })
            
            # 1.5 Store Visits (from import: same customer, different stores/dates - never miss an entry)
            for visit in client.visits.all().select_related('store').order_by('visit_date'):
                visit_date_iso = visit.visit_date.isoformat() if visit.visit_date else None
                journey_items.append({
                    'type': 'store_visit',
                    'id': visit.id,
                    'date': visit_date_iso,
                    'title': 'Store Visit',
                    'description': visit.store.name if visit.store else 'Store visit',
                    'details': {
                        'store': visit.store.name if visit.store else None,
                        'attended_by': visit.attended_by,
                    }
                })
            
            # 2. Client Interactions
            interactions = client.interactions.all().select_related('user')
            for interaction in interactions:
                store_name = interaction.user.store.name if interaction.user.store else "No Store"
                journey_items.append({
                    'type': 'interaction',
                    'id': interaction.id,
                    'date': interaction.created_at.isoformat() if interaction.created_at else None,
                    'title': f"{interaction.get_interaction_type_display()} - {interaction.subject}",
                    'description': interaction.description,
                    'details': {
                        'interaction_type': interaction.interaction_type,
                        'outcome': interaction.outcome,
                        'user': interaction.user.get_full_name() or interaction.user.username,
                        'store': store_name,
                    }
                })
            
            # 3. Appointments
            appointments = client.appointments.all().select_related('assigned_to', 'created_by', 'tenant')
            for appointment in appointments:
                store_name = "No Store"
                if appointment.assigned_to and appointment.assigned_to.store:
                    store_name = appointment.assigned_to.store.name
                elif appointment.created_by and appointment.created_by.store:
                    store_name = appointment.created_by.store.name
                
                appointment_datetime = timezone.make_aware(
                    timezone.datetime.combine(appointment.date, appointment.time)
                ) if appointment.date and appointment.time else None
                
                journey_items.append({
                    'type': 'appointment',
                    'id': appointment.id,
                    'date': appointment_datetime.isoformat() if appointment_datetime else (appointment.date.isoformat() if appointment.date else None),
                    'title': f"Appointment - {appointment.get_status_display()}",
                    'description': appointment.purpose,
                    'details': {
                        'status': appointment.status,
                        'purpose': appointment.purpose,
                        'location': appointment.location,
                        'assigned_to': appointment.assigned_to.get_full_name() if appointment.assigned_to else None,
                        'store': store_name,
                        'notes': appointment.notes,
                    }
                })
            
            # 4. Pipeline Entries
            pipelines = SalesPipeline.objects.filter(client=client).select_related('sales_representative')
            for pipeline in pipelines:
                store_name = pipeline.sales_representative.store.name if pipeline.sales_representative.store else "No Store"
                journey_items.append({
                    'type': 'pipeline',
                    'id': pipeline.id,
                    'date': pipeline.created_at.isoformat() if pipeline.created_at else None,
                    'title': f"Pipeline Entry - {pipeline.get_stage_display()}",
                    'description': pipeline.title,
                    'details': {
                        'stage': pipeline.stage,
                        'expected_value': float(pipeline.expected_value) if pipeline.expected_value else 0,
                        'probability': pipeline.probability,
                        'sales_rep': pipeline.sales_representative.get_full_name() or pipeline.sales_representative.username,
                        'store': store_name,
                        'notes': pipeline.notes,
                    }
                })
            
            # 5. Sales/Purchases
            sales = Sale.objects.filter(client=client).select_related('sales_representative')
            for sale in sales:
                store_name = sale.sales_representative.store.name if sale.sales_representative.store else "No Store"
                journey_items.append({
                    'type': 'sale',
                    'id': sale.id,
                    'date': sale.order_date.isoformat() if sale.order_date else (sale.created_at.isoformat() if sale.created_at else None),
                    'title': f"Purchase - {sale.get_status_display()}",
                    'description': f"Order #{sale.order_number}",
                    'details': {
                        'order_number': sale.order_number,
                        'total_amount': float(sale.total_amount) if sale.total_amount else 0,
                        'status': sale.status,
                        'payment_status': sale.payment_status,
                        'sales_rep': sale.sales_representative.get_full_name() or sale.sales_representative.username,
                        'store': store_name,
                    }
                })
            
            # 6. Follow-ups
            follow_ups = client.follow_ups.all().select_related('assigned_to', 'created_by')
            for follow_up in follow_ups:
                store_name = "No Store"
                if follow_up.assigned_to and follow_up.assigned_to.store:
                    store_name = follow_up.assigned_to.store.name
                elif follow_up.created_by and follow_up.created_by.store:
                    store_name = follow_up.created_by.store.name
                
                journey_items.append({
                    'type': 'followup',
                    'id': follow_up.id,
                    'date': follow_up.created_at.isoformat() if follow_up.created_at else None,
                    'title': f"Follow-up - {follow_up.get_status_display()}",
                    'description': follow_up.title,
                    'details': {
                        'status': follow_up.status,
                        'priority': follow_up.priority,
                        'due_date': follow_up.due_date.isoformat() if follow_up.due_date else None,
                        'assigned_to': follow_up.assigned_to.get_full_name() if follow_up.assigned_to else None,
                        'store': store_name,
                    }
                })
            
            # Sort by date (oldest first for timeline)
            journey_items.sort(key=lambda x: x['date'] or '1970-01-01', reverse=False)
            
            return Response({
                'success': True,
                'data': journey_items,
                'customer': {
                    'id': client.id,
                    'name': client.full_name,
                    'phone': client.phone,
                }
            })
            
        except Exception as e:
            import traceback
            print(f"Error fetching customer journey: {e}")
            print(traceback.format_exc())
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def mark_interest_purchased(self, request, pk=None, *args, **kwargs):
        """Mark a customer interest as purchased"""
        try:
            # When using as_view(), URL parameters are in self.kwargs
            # Extract pk and interest_id from URL path
            pk = pk or self.kwargs.get('pk')
            interest_id = self.kwargs.get('interest_id')
            
            if not pk:
                return Response(
                    {'error': 'Client ID is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not interest_id:
                return Response(
                    {'error': 'Interest ID is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Convert to int if they're strings
            try:
                pk = int(pk)
                interest_id = int(interest_id)
            except (ValueError, TypeError):
                return Response(
                    {'error': 'Invalid ID format'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get the client object
            client = Client.objects.get(pk=pk, tenant=request.user.tenant)
            interest = CustomerInterest.objects.get(id=interest_id, client=client, tenant=request.user.tenant)
            
            interest.is_purchased = True
            interest.is_not_purchased = False  # Reset not purchased if marking as purchased
            interest.purchased_at = timezone.now()
            interest.save()
            
            return Response({
                'success': True,
                'data': {
                    'id': interest.id,
                    'is_purchased': interest.is_purchased,
                    'is_not_purchased': interest.is_not_purchased,
                    'purchased_at': interest.purchased_at.isoformat() if interest.purchased_at else None
                },
                'message': f'Interest "{interest.product.name}" marked as purchased'
            })
        except CustomerInterest.DoesNotExist:
            return Response(
                {'error': 'Interest not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def mark_interest_not_purchased(self, request, pk=None, *args, **kwargs):
        """Mark a customer interest as not purchased"""
        try:
            # When using as_view(), URL parameters are in self.kwargs
            # Extract pk and interest_id from URL path
            pk = pk or self.kwargs.get('pk')
            interest_id = self.kwargs.get('interest_id')
            
            if not pk:
                return Response(
                    {'error': 'Client ID is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not interest_id:
                return Response(
                    {'error': 'Interest ID is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Convert to int if they're strings
            try:
                pk = int(pk)
                interest_id = int(interest_id)
            except (ValueError, TypeError):
                return Response(
                    {'error': 'Invalid ID format'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get the client object
            client = Client.objects.get(pk=pk, tenant=request.user.tenant)
            interest = CustomerInterest.objects.get(id=interest_id, client=client, tenant=request.user.tenant)
            
            interest.is_purchased = False
            interest.is_not_purchased = True
            interest.not_purchased_at = timezone.now()
            interest.save()
            
            return Response({
                'success': True,
                'data': {
                    'id': interest.id,
                    'is_purchased': interest.is_purchased,
                    'is_not_purchased': interest.is_not_purchased,
                    'not_purchased_at': interest.not_purchased_at.isoformat() if interest.not_purchased_at else None
                },
                'message': f'Interest "{interest.product.name}" marked as not purchased'
            })
        except CustomerInterest.DoesNotExist:
            return Response(
                {'error': 'Interest not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def dropdown_options(self, request):
        """Get dropdown options for customer form fields"""
        options = {
            'states': [
                {'value': 'AP', 'label': 'Andhra Pradesh'},
                {'value': 'AR', 'label': 'Arunachal Pradesh'},
                {'value': 'AS', 'label': 'Assam'},
                {'value': 'BR', 'label': 'Bihar'},
                {'value': 'CT', 'label': 'Chhattisgarh'},
                {'value': 'GA', 'label': 'Goa'},
                {'value': 'GJ', 'label': 'Gujarat'},
                {'value': 'HR', 'label': 'Haryana'},
                {'value': 'HP', 'label': 'Himachal Pradesh'},
                {'value': 'JK', 'label': 'Jammu and Kashmir'},
                {'value': 'JH', 'label': 'Jharkhand'},
                {'value': 'KA', 'label': 'Karnataka'},
                {'value': 'KL', 'label': 'Kerala'},
                {'value': 'MP', 'label': 'Madhya Pradesh'},
                {'value': 'MH', 'label': 'Maharashtra'},
                {'value': 'MN', 'label': 'Manipur'},
                {'value': 'ML', 'label': 'Meghalaya'},
                {'value': 'MZ', 'label': 'Mizoram'},
                {'value': 'NL', 'label': 'Nagaland'},
                {'value': 'OR', 'label': 'Odisha'},
                {'value': 'PB', 'label': 'Punjab'},
                {'value': 'RJ', 'label': 'Rajasthan'},
                {'value': 'SK', 'label': 'Sikkim'},
                {'value': 'TN', 'label': 'Tamil Nadu'},
                {'value': 'TG', 'label': 'Telangana'},
                {'value': 'TR', 'label': 'Tripura'},
                {'value': 'UP', 'label': 'Uttar Pradesh'},
                {'value': 'UT', 'label': 'Uttarakhand'},
                {'value': 'WB', 'label': 'West Bengal'},
                {'value': 'AN', 'label': 'Andaman and Nicobar Islands'},
                {'value': 'CH', 'label': 'Chandigarh'},
                {'value': 'DN', 'label': 'Dadra and Nagar Haveli'},
                {'value': 'DD', 'label': 'Daman and Diu'},
                {'value': 'DL', 'label': 'Delhi'},
                {'value': 'LD', 'label': 'Lakshadweep'},
                {'value': 'PY', 'label': 'Puducherry'},
            ],
            'communities': [
                {'value': 'hindu', 'label': 'Hindu'},
                {'value': 'muslim', 'label': 'Muslim'},
                {'value': 'sikh', 'label': 'Sikh'},
                {'value': 'christian', 'label': 'Christian'},
                {'value': 'jain', 'label': 'Jain'},
                {'value': 'buddhist', 'label': 'Buddhist'},
                {'value': 'parsi', 'label': 'Parsi'},
                {'value': 'jewish', 'label': 'Jewish'},
                {'value': 'gujarati', 'label': 'Gujarati'},
                {'value': 'marwari', 'label': 'Marwari'},
                {'value': 'punjabi', 'label': 'Punjabi'},
                {'value': 'sindhi', 'label': 'Sindhi'},
                {'value': 'bengali', 'label': 'Bengali'},
                {'value': 'tamil', 'label': 'Tamil'},
                {'value': 'telugu', 'label': 'Telugu'},
                {'value': 'kannada', 'label': 'Kannada'},
                {'value': 'malayalam', 'label': 'Malayalam'},
                {'value': 'marathi', 'label': 'Marathi'},
                {'value': 'hindi', 'label': 'Hindi'},
                {'value': 'urdu', 'label': 'Urdu'},
                {'value': 'kashmiri', 'label': 'Kashmiri'},
                {'value': 'assamese', 'label': 'Assamese'},
                {'value': 'oriya', 'label': 'Oriya'},
                {'value': 'other', 'label': 'Other'},
            ],
            'reasons_for_visit': [
                {'value': 'purchase', 'label': 'Purchase'},
                {'value': 'inquiry', 'label': 'Inquiry'},
                {'value': 'repair', 'label': 'Repair'},
                {'value': 'exchange', 'label': 'Exchange'},
                {'value': 'valuation', 'label': 'Valuation'},
                {'value': 'cleaning', 'label': 'Cleaning'},
                {'value': 'sizing', 'label': 'Sizing'},
                {'value': 'warranty', 'label': 'Warranty'},
                {'value': 'gift', 'label': 'Gift'},
                {'value': 'investment', 'label': 'Investment'},
                {'value': 'other', 'label': 'Other'},
            ],
            'lead_sources': [
                {'value': 'walkin', 'label': 'Walk-in'},
                {'value': 'referral', 'label': 'Referral'},
                {'value': 'online', 'label': 'Online'},
                {'value': 'social_media', 'label': 'Social Media'},
                {'value': 'advertisement', 'label': 'Advertisement'},
                {'value': 'exhibition', 'label': 'Exhibition'},
                {'value': 'cold_call', 'label': 'Cold Call'},
                {'value': 'website', 'label': 'Website'},
                {'value': 'google', 'label': 'Google Search'},
                {'value': 'facebook', 'label': 'Facebook'},
                {'value': 'instagram', 'label': 'Instagram'},
                {'value': 'whatsapp', 'label': 'WhatsApp'},
                {'value': 'newspaper', 'label': 'Newspaper'},
                {'value': 'magazine', 'label': 'Magazine'},
                {'value': 'tv', 'label': 'TV Advertisement'},
                {'value': 'radio', 'label': 'Radio Advertisement'},
                {'value': 'other', 'label': 'Other'},
            ],
            'age_groups': [
                {'value': '18-25', 'label': '18-25'},
                {'value': '26-35', 'label': '26-35'},
                {'value': '36-50', 'label': '36-50'},
                {'value': '51-65', 'label': '51-65'},
                {'value': '65+', 'label': '65+'},
            ],
            'saving_schemes': [
                {'value': 'active', 'label': 'Active'},
                {'value': 'inactive', 'label': 'Inactive'},
                {'value': 'pending', 'label': 'Pending'},
                {'value': 'completed', 'label': 'Completed'},
            ],
        }
        return Response(options)

class ClientInteractionViewSet(viewsets.ModelViewSet):
    queryset = ClientInteraction.objects.all()
    serializer_class = ClientInteractionSerializer
    permission_classes = [IsRoleAllowed.for_roles(['inhouse_sales', 'business_admin', 'manager'])]

class AppointmentViewSet(viewsets.ModelViewSet, ScopedVisibilityMixin):
    serializer_class = AppointmentSerializer
    permission_classes = [IsRoleAllowed.for_roles(['inhouse_sales', 'business_admin', 'manager'])]

    def list(self, request, *args, **kwargs):
        """List appointments with debugging"""
        print(f"=== APPOINTMENT LIST METHOD ===")
        print(f"Request user: {request.user}")
        print(f"Request method: {request.method}")
        print(f"Request URL: {request.path}")
        
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        response_data = serializer.data
        
        print(f"Serialized data count: {len(response_data)}")
        print(f"Serialized data: {response_data}")
        
        response = Response(response_data)
        print(f"=== APPOINTMENT LIST METHOD END ===")
        return response

    def get_queryset(self):
        """Filter appointments by user scope"""
        queryset = self.get_scoped_queryset(Appointment, is_deleted=False)
        
        # Optimize queries for serializer fields (client relationships)
        queryset = queryset.select_related(
            'client',
            'client__sales_person_id',
            'client__assigned_to',
            'client__created_by',
            'assigned_to',
            'created_by',
            'tenant'
        ).prefetch_related('client__interests__category', 'client__interests__product')
        
        # Exclude rescheduled appointments by default (they are replaced by new appointments)
        # Only include them if explicitly requested via status filter
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        else:
            # Default: exclude rescheduled appointments (use string value to ensure it works)
            queryset = queryset.exclude(status='rescheduled')
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if start_date:
            # Handle ISO datetime strings by extracting just the date part
            if 'T' in start_date:
                start_date = start_date.split('T')[0]
            queryset = queryset.filter(date__gte=start_date)
            
        if end_date:
            # Handle ISO datetime strings by extracting just the date part
            if 'T' in end_date:
                end_date = end_date.split('T')[0]
            queryset = queryset.filter(date__lte=end_date)
        
        # Filter by assigned user
        assigned_to = self.request.query_params.get('assigned_to')
        if assigned_to:
            queryset = queryset.filter(assigned_to_id=assigned_to)
        
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        tenant = user.tenant
        appointment = serializer.save(tenant=tenant, created_by=user, assigned_to=user)
        
        # Create notification for appointment creation
        self.create_appointment_notification(appointment, user)
        
        return appointment
    
    def create_appointment_notification(self, appointment, created_by_user):
        """Create notification when a new appointment is created."""
        try:
            from apps.notifications.models import Notification
            from apps.users.models import User
            
            print(f"=== CREATING APPOINTMENT NOTIFICATIONS ===")
            print(f"Appointment: {appointment.id}")
            print(f"Client: {appointment.client.first_name if appointment.client else 'No client'} {appointment.client.last_name if appointment.client else ''}")
            print(f"Client store: {appointment.client.store.name if appointment.client and appointment.client.store else 'No store'}")
            print(f"Created by user: {created_by_user.username} (role: {created_by_user.role})")
            print(f"Created by user tenant: {created_by_user.tenant}")
            
            # Get users to notify
            users_to_notify = []
            
            # The user who created the appointment should get notified
            users_to_notify.append(created_by_user)
            print(f"Added creator: {created_by_user.username}")
            
            # Notify the assigned user (if different from creator)
            if appointment.assigned_to and appointment.assigned_to != created_by_user:
                users_to_notify.append(appointment.assigned_to)
                print(f"Added assigned user: {appointment.assigned_to.username}")
            
            # Notify business admin
            if created_by_user.tenant:
                business_admins = User.objects.filter(
                    tenant=created_by_user.tenant,
                    role='business_admin',
                    is_active=True
                )
                print(f"Found {business_admins.count()} business admins: {[f'{admin.username} (active: {admin.is_active})' for admin in business_admins]}")
                users_to_notify.extend(business_admins)
            
            # Notify store users if appointment is for their store
            if appointment.client and appointment.client.store:
                print(f"Appointment has client with store: {appointment.client.store.name}")
                
                # Store manager
                store_managers = User.objects.filter(
                    tenant=created_by_user.tenant,
                    role='manager',
                    store=appointment.client.store,
                    is_active=True
                )
                store_managers_info = [f'{manager.username} (store: {manager.store.name if manager.store else "None"})' for manager in store_managers]
                print(f"Found {store_managers.count()} store managers: {store_managers_info}")
                users_to_notify.extend(store_managers)
                
                # In-house sales users
                inhouse_sales_users = User.objects.filter(
                    tenant=created_by_user.tenant,
                    role='inhouse_sales',
                    store=appointment.client.store,
                    is_active=True
                )
                inhouse_sales_info = [f'{user.username} (store: {user.store.name if user.store else "None"})' for user in inhouse_sales_users]
                print(f"Found {inhouse_sales_users.count()} inhouse sales users: {inhouse_sales_info}")
                users_to_notify.extend(inhouse_sales_users)
                
                # Tele-calling users
                telecalling_users = User.objects.filter(
                    tenant=created_by_user.tenant,
                    role='tele_calling',
                    store=appointment.client.store,
                    is_active=True
                )
                telecalling_info = [f'{user.username} (store: {user.store.name if user.store else "None"})' for user in telecalling_users]
                print(f"Found {telecalling_users.count()} telecalling users: {telecalling_info}")
                users_to_notify.extend(telecalling_users)
            else:
                print(f"Appointment has NO client or store - store users won't be notified")
                
                # If appointment has no client/store, notify all managers in the tenant
                all_managers = User.objects.filter(
                    tenant=created_by_user.tenant,
                    role='manager',
                    is_active=True
                )
                managers_info = [f'{manager.username} (store: {manager.store.name if manager.store else "None"})' for manager in all_managers]
                print(f"Found {all_managers.count()} managers in tenant (no store): {managers_info}")
                users_to_notify.extend(all_managers)
            
            # Remove duplicates
            unique_users = list({user.id: user for user in users_to_notify}.values())
            print(f"Total users to notify (before deduplication): {len(users_to_notify)}")
            print(f"Unique users to notify (after deduplication): {len(unique_users)}")
            
            # Create notifications
            for user in unique_users:
                notification = Notification.objects.create(
                    user=user,
                    tenant=appointment.tenant,
                    store=appointment.client.store if appointment.client else None,
                    type='appointment_reminder',
                    title='New appointment scheduled',
                    message=f'Appointment scheduled for {appointment.client.first_name} {appointment.client.last_name if appointment.client else "Customer"} on {appointment.date} at {appointment.time}',
                    priority='medium',
                    status='unread',
                    action_url=f'/appointments/{appointment.id}',
                    action_text='View Appointment',
                    is_persistent=False
                )
                print(f"Created notification {notification.id} for user {user.username} (role: {user.role})")
            
            print(f"Created {len(unique_users)} notifications for new appointment")
            print(f"Users notified: {[f'{user.username} ({user.role})' for user in unique_users]}")
            
        except Exception as e:
            print(f"Error creating appointment notification: {e}")
            import traceback
            print(f"Traceback: {traceback.format_exc()}")

    def perform_update(self, serializer):
        user = self.request.user
        serializer.save(updated_by=user)

    def perform_destroy(self, instance):
        # Soft delete
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.save()

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """Confirm an appointment"""
        appointment = self.get_object()
        appointment.status = Appointment.Status.CONFIRMED
        appointment.save()
        return Response({'status': 'confirmed'})

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Mark appointment as completed"""
        appointment = self.get_object()
        outcome_notes = request.data.get('outcome_notes')
        appointment.mark_completed(outcome_notes)
        return Response({'status': 'completed'})

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel an appointment"""
        appointment = self.get_object()
        reason = request.data.get('reason')
        appointment.cancel_appointment(reason)
        return Response({'status': 'cancelled'})

    @action(detail=True, methods=['post'])
    def reschedule(self, request, pk=None):
        """Reschedule an appointment"""
        appointment = self.get_object()
        new_date = request.data.get('new_date')
        new_time = request.data.get('new_time')
        reason = request.data.get('reason')
        
        if not new_date or not new_time:
            return Response(
                {'error': 'new_date and new_time are required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        new_appointment = appointment.reschedule_appointment(new_date, new_time, reason)
        return Response({
            'status': 'rescheduled',
            'new_appointment_id': new_appointment.id
        })

    @action(detail=False, methods=['get'])
    def slots(self, request):
        """Get available appointment slots for a given date range"""
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        # Get query parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        duration = int(request.query_params.get('duration', 60))  # Default 60 minutes
        
        if not start_date:
            start_date = timezone.now().date()
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            
        if not end_date:
            end_date = start_date + timedelta(days=7)  # Default to 1 week
        else:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Business hours (9 AM to 6 PM)
        business_start = datetime.strptime('09:00', '%H:%M').time()
        business_end = datetime.strptime('18:00', '%H:%M').time()
        
        # Generate time slots
        slots = []
        current_date = start_date
        
        while current_date <= end_date:
            # Skip weekends (Saturday = 5, Sunday = 6)
            if current_date.weekday() < 5:  # Monday to Friday
                current_time = business_start
                
                while current_time < business_end:
                    slot_end = (
                        datetime.combine(current_date, current_time) + 
                        timedelta(minutes=duration)
                    ).time()
                    
                    if slot_end <= business_end:
                        # Check if slot is available (no conflicting appointments)
                        conflicting_appointments = Appointment.objects.filter(
                            tenant=request.user.tenant,
                            date=current_date,
                            time__lt=slot_end,
                            time__gte=current_time,
                            status__in=['scheduled', 'confirmed'],
                            is_deleted=False
                        )
                        
                        if not conflicting_appointments.exists():
                            slots.append({
                                'date': current_date.strftime('%Y-%m-%d'),
                                'time': current_time.strftime('%H:%M'),
                                'end_time': slot_end.strftime('%H:%M'),
                                'duration': duration,
                                'available': True
                            })
                        else:
                            slots.append({
                                'date': current_date.strftime('%Y-%m-%d'),
                                'time': current_time.strftime('%H:%M'),
                                'end_time': slot_end.strftime('%H:%M'),
                                'duration': duration,
                                'available': False,
                                'conflicts': conflicting_appointments.count()
                            })
                    
                    # Move to next slot (30-minute intervals)
                    current_time = (
                        datetime.combine(current_date, current_time) + 
                        timedelta(minutes=30)
                    ).time()
            
            current_date += timedelta(days=1)
        
        return Response({
            'slots': slots,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'duration': duration,
            'business_hours': {
                'start': business_start.strftime('%H:%M'),
                'end': business_end.strftime('%H:%M')
            }
        })

    @action(detail=True, methods=['post'])
    def send_reminder(self, request, pk=None):
        """Send reminder for an appointment"""
        appointment = self.get_object()
        appointment.send_reminder()
        return Response({'status': 'reminder_sent'})

    @action(detail=False, methods=['get'])
    def calendar(self, request):
        """Get appointments for calendar view"""
        queryset = self.get_queryset()
        appointments = []
        
        for appointment in queryset:
            appointments.append({
                'id': appointment.id,
                'title': f"{appointment.client.full_name} - {appointment.purpose}",
                'start': f"{appointment.date}T{appointment.time}",
                'end': f"{appointment.date}T{appointment.time}",
                'status': appointment.status,
                'client_name': appointment.client.full_name,
                'purpose': appointment.purpose,
                'location': appointment.location,
                'assigned_to': appointment.assigned_to.get_full_name() if appointment.assigned_to else None,
            })
        
        return Response(appointments)

    # Debug endpoint removed for production security

    @action(detail=False, methods=['get'])
    def today(self, request):
        """Get today's appointments"""
        from django.utils import timezone
        today = timezone.now().date()
        queryset = self.get_queryset().filter(date=today)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        """Get upcoming appointments"""
        from django.utils import timezone
        today = timezone.now().date()
        queryset = self.get_queryset().filter(date__gte=today, status=Appointment.Status.SCHEDULED)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue appointments"""
        from django.utils import timezone
        today = timezone.now().date()
        queryset = self.get_queryset().filter(date__lt=today, status=Appointment.Status.SCHEDULED)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class FollowUpViewSet(viewsets.ModelViewSet, ScopedVisibilityMixin):
    serializer_class = FollowUpSerializer
    permission_classes = [IsRoleAllowed.for_roles(['inhouse_sales', 'manager', 'business_admin'])]

    def get_queryset(self):
        """Filter follow-ups by user scope"""
        queryset = self.get_scoped_queryset(FollowUp, is_deleted=False)
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by priority
        priority_filter = self.request.query_params.get('priority')
        if priority_filter:
            queryset = queryset.filter(priority=priority_filter)
        
        # Filter by type
        type_filter = self.request.query_params.get('type')
        if type_filter:
            queryset = queryset.filter(type=type_filter)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(due_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(due_date__lte=end_date)
        
        # Filter by assigned user
        assigned_to = self.request.query_params.get('assigned_to')
        if assigned_to:
            queryset = queryset.filter(assigned_to_id=assigned_to)
        
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        tenant = user.tenant
        serializer.save(tenant=tenant, created_by=user, assigned_to=user)

    def perform_update(self, serializer):
        user = self.request.user
        serializer.save(updated_by=user)

    def perform_destroy(self, instance):
        # Soft delete
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.save()

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Mark follow-up as completed"""
        follow_up = self.get_object()
        outcome_notes = request.data.get('outcome_notes')
        follow_up.mark_completed(outcome_notes)
        return Response({'status': 'completed'})

    @action(detail=True, methods=['post'])
    def send_reminder(self, request, pk=None):
        """Send reminder for a follow-up"""
        follow_up = self.get_object()
        follow_up.send_reminder()
        return Response({'status': 'reminder_sent'})

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue follow-ups"""
        from django.utils import timezone
        today = timezone.now().date()
        queryset = self.get_queryset().filter(due_date__lt=today, status=FollowUp.Status.PENDING)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def due_today(self, request):
        """Get follow-ups due today"""
        from django.utils import timezone
        today = timezone.now().date()
        queryset = self.get_queryset().filter(due_date=today, status=FollowUp.Status.PENDING)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        """Get upcoming follow-ups"""
        from django.utils import timezone
        today = timezone.now().date()
        queryset = self.get_queryset().filter(due_date__gte=today, status=FollowUp.Status.PENDING)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TaskViewSet(viewsets.ModelViewSet, ScopedVisibilityMixin):
    serializer_class = TaskSerializer
    permission_classes = [IsRoleAllowed.for_roles(['inhouse_sales', 'manager', 'business_admin'])]

    def get_queryset(self):
        return self.get_scoped_queryset(Task)

    def perform_create(self, serializer):
        user = self.request.user
        # Debug statements removed for production
        tenant = user.tenant
        serializer.save(tenant=tenant, created_by=user, assigned_to=user)

class AnnouncementViewSet(viewsets.ModelViewSet):
    queryset = Announcement.objects.all()
    serializer_class = AnnouncementSerializer
    permission_classes = [IsRoleAllowed.for_roles(['inhouse_sales', 'manager', 'business_admin'])]

class PurchaseViewSet(viewsets.ModelViewSet):
    queryset = Purchase.objects.all()
    serializer_class = PurchaseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Purchase.objects.all()
        user = self.request.user
        if user.is_authenticated and user.tenant:
            queryset = queryset.filter(client__tenant=user.tenant)
        else:
            queryset = Purchase.objects.none()
        client_id = self.request.query_params.get('client')
        if client_id:
            queryset = queryset.filter(client_id=client_id)
        return queryset

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = AuditLogSerializer
    permission_classes = [IsRoleAllowed.for_roles(['inhouse_sales','business_admin','manager'])]

    def get_queryset(self):
        queryset = AuditLog.objects.all().order_by('-timestamp')
        client_id = self.request.query_params.get('client')
        user = self.request.user
        if user.is_authenticated and user.is_manager:
            # Managers see audit logs for customers in their tenant only
            if user.tenant:
                queryset = queryset.filter(client__tenant=user.tenant)
            else:
                queryset = AuditLog.objects.none()
        if client_id:
            queryset = queryset.filter(client_id=client_id)
        return queryset


class CustomerTagViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for CustomerTag model - provides read-only access to customer tags
    """
    serializer_class = CustomerTagSerializer
    permission_classes = [IsRoleAllowed.for_roles(['inhouse_sales', 'manager', 'business_admin', 'tele_calling', 'marketing'])]
    
    def get_queryset(self):
        """Return all active customer tags"""
        return CustomerTag.objects.filter(is_active=True)
    
    @action(detail=False, methods=['get'])
    def by_category(self, request):
        """Get tags grouped by category"""
        category = request.query_params.get('category')
        queryset = self.get_queryset()
        
        if category:
            queryset = queryset.filter(category=category)
        
        # Group by category
        categories = {}
        for tag in queryset:
            if tag.category not in categories:
                categories[tag.category] = []
            categories[tag.category].append({
                'id': tag.id,
                'name': tag.name,
                'slug': tag.slug,
                'category': tag.category,
                'description': tag.description
            })
        
        return Response(categories)
    
    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Get available tag categories"""
        categories = CustomerTag.CATEGORY_CHOICES
        return Response([{'value': choice[0], 'label': choice[1]} for choice in categories])
